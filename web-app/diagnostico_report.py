"""Diagnóstico Financiero: generación del reporte Excel y envío por correo.

Los datos vienen del formulario público /diagnostico/ (sin login). El payload
sanitizado se persiste en asesoria_db (base SEPARADA de la de Neto app — ver
asesoria_schema.sql; lo hace run.py vía db.save_diagnostico) para los reportes
del servicio de asesoría, y además se arma un .xlsx en memoria que se envía
por SMTP (neto@investorcr.com) al cliente y al asesor.

Variables de entorno requeridas (.env de la raíz, ya presentes en el server):
  SMTP_HOST / SMTP_PORT / SMTP_USER / SMTP_PASSWORD
"""

import html as html_mod
import io
import os
import re
import smtplib
from datetime import datetime
from email.message import EmailMessage
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ADVISOR_EMAIL = os.environ.get("DIAGNOSTICO_ADVISOR_EMAIL", "dylanmos96@gmail.com")

_CR_TZ = ZoneInfo("America/Costa_Rica")
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# Límites de sanitización (el endpoint es público)
_MAX_ROWS = 100
_MAX_TEXT = 300
_MAX_NOTAS = 2000

# Estilos neto
_INK = "1B1C20"
_TEAL = "3A9C8E"
_FILL_SECTION = PatternFill("solid", fgColor=_INK)
_FILL_SUBTOTAL = PatternFill("solid", fgColor="E2E6EB")
_FONT_TITLE = Font(name="Calibri", size=14, bold=True, color=_INK)
_FONT_SECTION = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_FONT_BOLD = Font(name="Calibri", size=11, bold=True)
_CRC_FMT = '"₡"#,##0'
_THIN = Border(bottom=Side(style="thin", color="E2E6EB"))


def _clean_text(value, max_len=_MAX_TEXT):
    return str(value or "").strip()[:max_len]


def _clean_amount(value):
    try:
        n = float(value)
    except (TypeError, ValueError):
        return 0.0
    if n != n or n in (float("inf"), float("-inf")) or n < 0:
        return 0.0
    return min(n, 1e12)


# ── Personalidades del dinero (Olivia Mellan) ────────────────────────────────

_ORDEN_PERFILES = ["ahorrador", "gastador", "evasor", "monje", "amasador", "preocupado"]

PERFILES_DEF = {
    "ahorrador": ("🏦", "El Ahorrador",
                  "Su disciplina para guardar es una fortaleza que pocos tienen — la "
                  "seguridad es su motor. Su riesgo: vivir con una escasez autoimpuesta "
                  "y postergar indefinidamente el disfrute del fruto de su trabajo."),
    "gastador": ("🛍️", "El Gastador",
                 "Sabe disfrutar la vida y ser generoso con los suyos — el dinero fluye "
                 "y le da alegría. Su riesgo: poca capacidad de ahorro, compras "
                 "impulsivas y deudas de consumo que hipotecan ese mismo disfrute futuro."),
    "evasor": ("🙈", "El Evasor",
               "Suele ser una persona capaz y ocupada en lo que sí le importa — el tema "
               "no es falta de habilidad, sino que el dinero le genera ansiedad y lo "
               "pospone. Su riesgo: desorden, cobros vencidos y sorpresas desagradables "
               "que se pudieron evitar."),
    "monje": ("🕊️", "El Monje del dinero",
              "Tiene valores sólidos y claridad de que el dinero no es lo más importante "
              "en la vida. Su riesgo: creer que quererlo es \"malo\" puede llevarlo a "
              "sabotear sus propios ingresos y oportunidades sin darse cuenta."),
    "amasador": ("📈", "El Amasador",
                 "Tiene una capacidad natural para hacer crecer el patrimonio — el dinero "
                 "es logro, libertad y opciones. Su riesgo: que nunca sea suficiente, con "
                 "exceso de trabajo y descuido de relaciones y salud en el camino."),
    "preocupado": ("😰", "El Preocupado",
                   "Es prevenido y responsable — nada lo toma por sorpresa porque siempre "
                   "está atento. Su riesgo: que el dinero sea una fuente permanente de "
                   "estrés aunque los números estén bien, revisando compulsivamente sin "
                   "ganar tranquilidad."),
}

_MIN_RESPUESTAS = 5


def clasificar_personalidad(respuestas):
    """Puntuación Mellan: 1 punto al perfil de cada respuesta. Devuelve
    {predominante, secundario, n} o None si hay menos de 5 respuestas.
    El secundario solo se reporta si acumula 2 o más puntos."""
    votos = [r for r in respuestas if r]
    if len(votos) < _MIN_RESPUESTAS:
        return None
    conteo = {k: votos.count(k) for k in _ORDEN_PERFILES}
    predominante = max(_ORDEN_PERFILES, key=lambda k: conteo[k])
    resto = [k for k in _ORDEN_PERFILES if k != predominante]
    secundario = max(resto, key=lambda k: conteo[k])
    if conteo[secundario] < 2:
        secundario = None
    return {"predominante": predominante, "secundario": secundario, "n": len(votos)}


def _clean_personalidades(raw):
    """Normaliza [{nombre, respuestas[8]}] (máx. 2 personas, valores del quiz)."""
    personas = []
    if not isinstance(raw, list):
        return personas
    for p in raw[:2]:
        if not isinstance(p, dict):
            continue
        nombre = _clean_text(p.get("nombre"), 120)
        respuestas = []
        resp_raw = p.get("respuestas")
        if isinstance(resp_raw, list):
            for r in resp_raw[:len(_ORDEN_PERFILES) + 2]:
                r = str(r or "").strip().lower()
                respuestas.append(r if r in _ORDEN_PERFILES else None)
        respuestas = respuestas[:8]
        if nombre or any(respuestas):
            personas.append({"nombre": nombre, "respuestas": respuestas})
    return personas


def _clean_currency(value):
    cur = str(value or "crc").strip().lower()
    return cur if cur in ("crc", "usd") else "crc"


def _clean_rows(raw, fields=("amount",)):
    """Normaliza una lista [{name, <montos>, currency}] recortando filas y valores."""
    rows = []
    if not isinstance(raw, list):
        return rows
    for item in raw[:_MAX_ROWS]:
        if not isinstance(item, dict):
            continue
        row = {"name": _clean_text(item.get("name")),
               "currency": _clean_currency(item.get("currency"))}
        for f in fields:
            row[f] = _clean_amount(item.get(f))
        if row["name"] or any(row[f] for f in fields):
            rows.append(row)
    return rows


def sanitize_payload(data):
    """Valida y normaliza el JSON del formulario. Devuelve (payload, error)."""
    if not isinstance(data, dict):
        return None, "Datos inválidos."

    nombre = _clean_text(data.get("nombre"), 120)
    correo = _clean_text(data.get("correo"), 200).lower()
    celular = _clean_text(data.get("celular"), 40)

    if len(nombre) < 3:
        return None, "Ingrese su nombre completo."
    if not _EMAIL_RE.match(correo):
        return None, "Ingrese un correo electrónico válido."
    if len(re.sub(r"\D", "", celular)) < 8:
        return None, "Ingrese un número celular válido (mínimo 8 dígitos)."

    payload = {
        "nombre": nombre,
        "correo": correo,
        "celular": celular,
        "fijos": _clean_rows(data.get("fijos")),
        "variables": _clean_rows(data.get("variables")),
        "hormiga": _clean_rows(data.get("hormiga")),
        "ingresos": _clean_rows(data.get("ingresos")),
        "activos": _clean_rows(data.get("activos")),
        "deudas": _clean_rows(data.get("deudas"),
                              fields=("saldo", "tasa", "cuota", "plazo")),
        "fe_tiene": _clean_text(data.get("fe_tiene"), 10),
        "fe_monto": _clean_amount(data.get("fe_monto")),
        "fe_moneda": _clean_currency(data.get("fe_moneda")),
        "seguros": [_clean_text(s, 60) for s in (data.get("seguros") or [])[:20]
                    if _clean_text(s, 60)],
        "retiro": _clean_text(data.get("retiro"), 120),
        "metas": [_clean_text(m) for m in (data.get("metas") or [])[:3]],
        "notas": _clean_text(data.get("notas"), _MAX_NOTAS),
        "personalidades": _clean_personalidades(data.get("personalidades")),
    }
    # El plazo es un conteo de meses, no un monto: entero y tope de 50 años.
    for d in payload["deudas"]:
        d["plazo"] = min(int(round(d["plazo"])), 600)
    return payload, None


# ── Conversión USD → CRC ─────────────────────────────────────────────────────
# Los inputs del formulario aceptan colones o dólares; todos los agregados y
# el reporte van en colones. La conversión usa el tipo de cambio más reciente
# de core.exchange_rates (lo obtiene run.py vía db.compute_amount_local).

_CATS_MONTOS = ("fijos", "variables", "hormiga", "ingresos", "activos")


def usa_usd(p):
    """True si el payload trae algún monto en dólares."""
    if p.get("fe_moneda") == "usd" and p.get("fe_monto"):
        return True
    for cat in _CATS_MONTOS:
        if any(r.get("currency") == "usd" for r in p[cat]):
            return True
    return any(d.get("currency") == "usd" for d in p["deudas"])


def aplicar_tipo_cambio(p, tc_usd, tc_fecha=None):
    """Convierte in-place los montos en USD a colones y anota el monto original
    en la descripción (transparencia para el asesor). Deja constancia del tipo
    de cambio usado en p['tc_usd'] / p['tc_fecha'] para el Excel."""
    for cat in _CATS_MONTOS:
        for r in p[cat]:
            if r.get("currency") == "usd":
                if r["amount"]:
                    r["name"] = ((r["name"] or "(sin descripción)")
                                 + f" [US$ {r['amount']:,.2f}]")
                    r["amount"] = round(r["amount"] * tc_usd, 2)
                r["currency"] = "crc"
    for d in p["deudas"]:
        if d.get("currency") == "usd":
            if d["saldo"] or d["cuota"]:
                d["name"] = ((d["name"] or "(sin nombre)")
                             + f" [US$: saldo {d['saldo']:,.2f}, cuota {d['cuota']:,.2f}]")
                d["saldo"] = round(d["saldo"] * tc_usd, 2)
                d["cuota"] = round(d["cuota"] * tc_usd, 2)
            d["currency"] = "crc"
    if p.get("fe_moneda") == "usd":
        p["fe_monto"] = round(p["fe_monto"] * tc_usd, 2)
        p["fe_moneda"] = "crc"
    p["tc_usd"] = float(tc_usd)
    p["tc_fecha"] = str(tc_fecha) if tc_fecha else None
    return p


# ── Clasificación general (flujo × patrimonio) ──────────────────────────────
# Escala de 4 niveles expresada en positivo, de mayor rango de mejora a más
# estable. El nivel 2 tiene dos variantes PARALELAS (no consecutivas): tener
# una sola de las dos columnas en pie — el patrimonio (Reconstrucción) o el
# flujo (Impulso). Nadie pasa de una variante a la otra: de ambas se salta al
# nivel 3 al tener las dos en positivo.
# Bordes: flujo ≥ 0 y patrimonio ≥ 0 cuentan como positivos; 20% = regla 50-30-20.

NIVELES_CLASIFICACION = [
    (1, "🌱", "Reorganización",
     "Dar este paso y ver los números completos ya es el punto de partida. El "
     "mayor rango de mejora está en el flujo mensual: pequeños ajustes liberan "
     "efectivo rápido, y de ahí se construye todo lo demás."),
    (2, "🧱", "Reconstrucción",
     "Hay un patrimonio construido que respalda — eso es un logro real. El "
     "reto ahora es que el mes a mes no lo erosione: recuperar un flujo "
     "positivo es la forma de proteger lo ya construido."),
    (2, "🚀", "Impulso",
     "La dirección es la correcta: cada mes se genera excedente. Sostener ese "
     "ritmo irá reduciendo las deudas hasta darle vuelta al patrimonio — es "
     "cuestión de constancia, y la trayectoria es lo que más pesa."),
    (3, "⚖️", "Estabilidad",
     "Flujo positivo y patrimonio a favor: la base está sólida. El siguiente "
     "nivel es llevar el excedente hacia el 20% del ingreso (la regla "
     "50-30-20) para acelerar las metas."),
    (4, "🏆", "Libertad financiera",
     "Se vive la regla 50-30-20: se ahorra el 20% o más del ingreso con "
     "patrimonio positivo. La conversación ya no es de orden, sino de "
     "crecimiento: inversión y metas de largo plazo."),
]

NIVEL_MAXIMO = 4


def clasificacion_general(t):
    """{nivel 1-4, total 4, emoji, nombre, desc} o None si no hay datos."""
    if not (t["ingresos"] > 0 or t["gastos"] > 0 or t["activos"] > 0 or t["deudas"] > 0):
        return None
    fc_pos = t["flujo"] >= 0
    pn_pos = t["patrimonio"] >= 0
    if not fc_pos and not pn_pos:
        idx = 0  # Reorganización
    elif not fc_pos and pn_pos:
        idx = 1  # Reconstrucción (nivel 2, variante patrimonio)
    elif fc_pos and not pn_pos:
        idx = 2  # Impulso (nivel 2, variante flujo)
    elif t["ingresos"] > 0 and t["flujo"] / t["ingresos"] >= 0.20:
        idx = 4  # Libertad financiera
    else:
        idx = 3  # Estabilidad
    nivel, emoji, nombre, desc = NIVELES_CLASIFICACION[idx]
    return {"nivel": nivel, "total": NIVEL_MAXIMO, "emoji": emoji,
            "nombre": nombre, "desc": desc}


def _totales(p):
    tf = sum(r["amount"] for r in p["fijos"])
    tv = sum(r["amount"] for r in p["variables"])
    th = sum(r["amount"] for r in p["hormiga"])
    ti = sum(r["amount"] for r in p["ingresos"])
    ta = sum(r["amount"] for r in p["activos"])
    td = sum(r["saldo"] for r in p["deudas"])
    tc = sum(r["cuota"] for r in p["deudas"])
    gastos = tf + tv + th + tc
    return {
        "fijos": tf, "variables": tv, "hormiga": th, "ingresos": ti,
        "activos": ta, "deudas": td, "cuotas": tc, "gastos": gastos,
        "flujo": ti - gastos, "patrimonio": ta - td,
    }


def build_excel(p):
    """Arma el .xlsx del diagnóstico en memoria y devuelve los bytes."""
    t = _totales(p)
    fecha = datetime.now(_CR_TZ)

    wb = Workbook()
    ws = wb.active
    ws.title = "Diagnóstico"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.sheet_view.showGridLines = False

    row = 1

    def put(col, r, value, font=None, fmt=None, fill=None, align=None):
        cell = ws.cell(row=r, column=col, value=value)
        if font:
            cell.font = font
        if fmt:
            cell.number_format = fmt
        if fill:
            cell.fill = fill
        if align:
            cell.alignment = Alignment(horizontal=align)
        return cell

    def section(title):
        nonlocal row
        row += 1
        for c in range(1, 6):
            put(c, row, None, fill=_FILL_SECTION)
        put(1, row, title, font=_FONT_SECTION, fill=_FILL_SECTION)
        row += 1

    def item(label, value, fmt=_CRC_FMT, bold=False, fill=None):
        nonlocal row
        put(1, row, label, font=_FONT_BOLD if bold else None, fill=fill)
        put(2, row, value, font=_FONT_BOLD if bold else None, fmt=fmt, fill=fill,
            align="right")
        row += 1

    put(1, row, "DIAGNÓSTICO FINANCIERO PERSONAL — neto", font=_FONT_TITLE)
    row += 1
    put(1, row, "Fecha: " + fecha.strftime("%d/%m/%Y %H:%M") + " (hora Costa Rica)")
    row += 1

    section("DATOS DEL CLIENTE")
    item("Nombre completo", p["nombre"], fmt="General")
    item("Correo electrónico", p["correo"], fmt="General")
    item("Número celular", p["celular"], fmt="General")
    if p.get("tc_usd"):
        fecha_tc = f" ({p['tc_fecha']})" if p.get("tc_fecha") else ""
        item("Tipo de cambio aplicado",
             f"US$ 1 = ₡{p['tc_usd']:,.2f}{fecha_tc}", fmt="General")

    c = clasificacion_general(t)
    if c:
        section("CLASIFICACIÓN GENERAL")
        item("Nivel", f"{c['emoji']} Nivel {c['nivel']} de {c['total']} — {c['nombre']}",
             fmt="General", bold=True)
        item("Lectura", c["desc"], fmt="General")

    section("RESUMEN — FLUJO MENSUAL")
    item("Ingresos", t["ingresos"])
    item("Gastos fijos", t["fijos"])
    item("Gastos variables", t["variables"])
    item("Gastos no necesarios", t["hormiga"])
    item("Cuotas de deudas", t["cuotas"])
    item("FLUJO DE CAJA", t["flujo"], bold=True, fill=_FILL_SUBTOTAL)
    if t["ingresos"] > 0:
        item("Carga de deuda (cuotas / ingreso)", t["cuotas"] / t["ingresos"],
             fmt="0.0%")
        item("Gastos no necesarios (% del ingreso)", t["hormiga"] / t["ingresos"],
             fmt="0.0%")
        item("Tasa de ahorro (flujo / ingreso)", t["flujo"] / t["ingresos"],
             fmt="0.0%")

    section("RESUMEN — PATRIMONIO")
    item("Activos", t["activos"])
    item("Deudas (saldo)", t["deudas"])
    item("PATRIMONIO NETO", t["patrimonio"], bold=True, fill=_FILL_SUBTOTAL)

    def detalle(titulo, rows, total):
        section(titulo)
        for r in rows:
            item(r["name"] or "(sin descripción)", r["amount"])
        item("Subtotal", total, bold=True, fill=_FILL_SUBTOTAL)

    detalle("GASTOS FIJOS", p["fijos"], t["fijos"])
    detalle("GASTOS VARIABLES", p["variables"], t["variables"])
    detalle("GASTOS NO NECESARIOS", p["hormiga"], t["hormiga"])
    detalle("INGRESOS MENSUALES", p["ingresos"], t["ingresos"])
    detalle("ACTIVOS", p["activos"], t["activos"])

    section("DEUDAS")
    put(1, row, "Deuda", font=_FONT_BOLD)
    put(2, row, "Saldo total", font=_FONT_BOLD, align="right")
    put(3, row, "Tasa anual", font=_FONT_BOLD, align="right")
    put(4, row, "Cuota mensual", font=_FONT_BOLD, align="right")
    put(5, row, "Plazo (meses)", font=_FONT_BOLD, align="right")
    row += 1
    for d in p["deudas"]:
        put(1, row, d["name"] or "(sin nombre)")
        put(2, row, d["saldo"], fmt=_CRC_FMT, align="right")
        put(3, row, d["tasa"] / 100.0, fmt="0.0%", align="right")
        put(4, row, d["cuota"], fmt=_CRC_FMT, align="right")
        # .get: los payloads guardados antes de agregar el campo no lo traen.
        put(5, row, d.get("plazo") or None, fmt="0", align="right")
        row += 1
    put(1, row, "Total", font=_FONT_BOLD, fill=_FILL_SUBTOTAL)
    put(2, row, t["deudas"], font=_FONT_BOLD, fmt=_CRC_FMT, fill=_FILL_SUBTOTAL,
        align="right")
    put(3, row, None, fill=_FILL_SUBTOTAL)
    put(4, row, t["cuotas"], font=_FONT_BOLD, fmt=_CRC_FMT, fill=_FILL_SUBTOTAL,
        align="right")
    put(5, row, None, fill=_FILL_SUBTOTAL)
    row += 1

    section("PROTECCIÓN Y RETIRO")
    item("¿Tiene fondo de emergencia?",
         {"si": "Sí", "no": "No"}.get(p["fe_tiene"], "—"), fmt="General")
    item("Monto disponible para emergencias", p["fe_monto"])
    gastos_basicos = t["fijos"] + t["variables"]
    if gastos_basicos > 0:
        item("Meses de gastos cubiertos (meta 3–6)",
             p["fe_monto"] / gastos_basicos, fmt="0.0")
    item("Seguros vigentes", ", ".join(p["seguros"]) or "Ninguno reportado",
         fmt="General")
    item("Retiro", p["retiro"] or "—", fmt="General")

    section("METAS FINANCIERAS")
    for i, m in enumerate(p["metas"], start=1):
        if m:
            item(f"Meta {i}", m, fmt="General")
    if p["notas"]:
        item("Notas para el asesor", p["notas"], fmt="General")

    if p.get("personalidades"):
        section("PERSONALIDADES DEL DINERO (OLIVIA MELLAN)")
        for i, persona in enumerate(p["personalidades"], start=1):
            nombre = persona["nombre"] or f"Persona {i}"
            res = clasificar_personalidad(persona["respuestas"])
            if res is None:
                n = len([r for r in persona["respuestas"] if r])
                item(nombre, f"Respuestas insuficientes ({n} de 8; mínimo {_MIN_RESPUESTAS})",
                     fmt="General")
                continue
            emoji, titulo, _ = PERFILES_DEF[res["predominante"]]
            item(f"{nombre} — perfil predominante", f"{emoji} {titulo}", fmt="General")
            if res["secundario"]:
                emoji2, titulo2, _ = PERFILES_DEF[res["secundario"]]
                item(f"{nombre} — rasgo secundario", f"{emoji2} {titulo2}", fmt="General")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_LOGO_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "static", "rebranding", "logos", "neto-logo-transparent-600w.png",
)

# Barra de proporciones 50-30-20 con la paleta del favicon (donut):
# teal (50%) → azul (30%) → amarillo (20%)
_BAR_5030_20 = f"""
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0">
  <tr>
    <td width="50%" height="8" bgcolor="#3A9C8E" style="font-size:0;line-height:0;">&nbsp;</td>
    <td width="30%" height="8" bgcolor="#1F4D8E" style="font-size:0;line-height:0;">&nbsp;</td>
    <td width="20%" height="8" bgcolor="#EFA91A" style="font-size:0;line-height:0;">&nbsp;</td>
  </tr>
</table>"""


def _fmt_crc(n):
    """₡1.234.567 (separador de miles al estilo es-CR)."""
    return "₡" + f"{round(n):,}".replace(",", ".")


def _resultados_personalidades(p):
    """[(nombre, resultado_o_None, n_respondidas)] por persona del quiz."""
    out = []
    for i, persona in enumerate(p.get("personalidades") or [], start=1):
        nombre = persona["nombre"] or f"Persona {i}"
        n = len([r for r in persona["respuestas"] if r])
        out.append((nombre, clasificar_personalidad(persona["respuestas"]), n))
    return out


def _html_personalidades(p):
    """Bloque HTML del correo con los perfiles de personalidad (o '' si no hay quiz)."""
    esc = html_mod.escape
    resultados = _resultados_personalidades(p)
    if not resultados:
        return ""

    def tarjeta(etiqueta, perfil, color):
        emoji, titulo, definicion = PERFILES_DEF[perfil]
        return f"""
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0"
             style="margin-top:10px;">
        <tr>
          <td width="4" bgcolor="{color}" style="border-radius:4px 0 0 4px;font-size:0;">&nbsp;</td>
          <td bgcolor="#F4F6F9" style="padding:14px 16px;border-radius:0 4px 4px 0;
              font-family:Helvetica,Arial,sans-serif;">
            <div style="font-size:10px;font-weight:bold;letter-spacing:.06em;
                        text-transform:uppercase;color:#6B7280;">{etiqueta}</div>
            <div style="font-size:15px;font-weight:bold;color:#1B1C20;padding-top:3px;">
              {emoji} {esc(titulo)}</div>
            <div style="font-size:12.5px;line-height:1.6;color:#4B5563;padding-top:4px;">
              {esc(definicion)}</div>
          </td>
        </tr>
      </table>"""

    bloques = []
    for nombre, res, n in resultados:
        if res is None:
            cuerpo = f"""
      <p style="font-size:12.5px;line-height:1.6;color:#6B7280;margin:8px 0 0;">
        Respondió {n} de 8 preguntas — se necesitan al menos {_MIN_RESPUESTAS} para
        identificar el perfil. Lo pueden completar juntos en la sesión.</p>"""
        else:
            cuerpo = tarjeta("Perfil predominante", res["predominante"], "#3A9C8E")
            if res["secundario"]:
                cuerpo += tarjeta("Rasgo secundario", res["secundario"], "#EFA91A")
        bloques.append(f"""
      <div style="font-size:15px;font-weight:bold;letter-spacing:-.01em;margin-top:22px;">
        {esc(nombre)}</div>{cuerpo}""")

    return f"""
      <div style="font-size:17px;font-weight:bold;letter-spacing:-.02em;margin-top:32px;">
        🧭 Personalidades del dinero</div>
      <p style="font-size:12.5px;line-height:1.6;color:#6B7280;margin:6px 0 0;">
        Basado en las clasificaciones de Olivia Mellan. Conocer su relación con el
        dinero ayuda a diseñar un plan que sí se pueda sostener.</p>
      {''.join(bloques)}"""


def _texto_clasificacion(p):
    """Versión de texto plano de la clasificación general ('' si no hay datos)."""
    c = clasificacion_general(_totales(p))
    if not c:
        return ""
    return ("\n— CLASIFICACIÓN GENERAL —\n"
            f"Nivel {c['nivel']} de {c['total']} — {c['nombre']}. {c['desc']}\n")


def _texto_personalidades(p):
    """Versión de texto plano del bloque de personalidades ('' si no hay quiz)."""
    resultados = _resultados_personalidades(p)
    if not resultados:
        return ""
    lineas = ["", "— PERSONALIDADES DEL DINERO —"]
    for nombre, res, n in resultados:
        if res is None:
            lineas.append(f"{nombre}: respuestas insuficientes ({n} de 8; mínimo {_MIN_RESPUESTAS}).")
            continue
        emoji, titulo, definicion = PERFILES_DEF[res["predominante"]]
        lineas.append(f"{nombre} — Perfil predominante: {titulo}. {definicion}")
        if res["secundario"]:
            emoji2, titulo2, definicion2 = PERFILES_DEF[res["secundario"]]
            lineas.append(f"{nombre} — Rasgo secundario: {titulo2}. {definicion2}")
    return "\n".join(lineas) + "\n"


def _html_clasificacion(t):
    """Banner HTML de la clasificación general ('' si no hay datos)."""
    esc = html_mod.escape
    c = clasificacion_general(t)
    if not c:
        return ""
    separador = '<td width="5" style="font-size:0;">&nbsp;</td>'
    partes = []
    for i in range(1, NIVEL_MAXIMO + 1):
        color = "#3A9C8E" if i <= c["nivel"] else "#E2E6EB"
        partes.append(f'<td height="8" bgcolor="{color}" '
                      'style="border-radius:4px;font-size:0;line-height:0;">&nbsp;</td>')
        if i < NIVEL_MAXIMO:
            partes.append(separador)
    segmentos = "".join(partes)
    return f"""
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0"
             style="margin:0 0 24px;">
        <tr>
          <td width="4" bgcolor="#3A9C8E" style="border-radius:4px 0 0 4px;font-size:0;">&nbsp;</td>
          <td bgcolor="#F0F7F6" style="padding:16px 18px;border-radius:0 4px 4px 0;
              font-family:Helvetica,Arial,sans-serif;">
            <div style="font-size:10px;font-weight:bold;letter-spacing:.06em;
                        text-transform:uppercase;color:#6B7280;">Clasificación general</div>
            <div style="font-size:18px;font-weight:bold;color:#1B1C20;padding-top:4px;">
              {c["emoji"]} Nivel {c["nivel"]} de {c["total"]} — {esc(c["nombre"])}</div>
            <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0"
                   style="margin-top:10px;"><tr>{segmentos}</tr></table>
            <div style="font-size:12.5px;line-height:1.6;color:#4B5563;padding-top:10px;">
              {esc(c["desc"])}</div>
          </td>
        </tr>
      </table>"""


def _build_html(p, fecha_larga, archivo):
    """Cuerpo HTML del correo con la marca neto (logo inline cid:netologo)."""
    esc = html_mod.escape
    t = _totales(p)
    flujo_color = "#3A9C8E" if t["flujo"] >= 0 else "#B3402E"
    patri_color = "#3A9C8E" if t["patrimonio"] >= 0 else "#B3402E"

    def kpi(label, value, color="#1B1C20"):
        return f"""
        <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0"
               style="background:#F4F6F9;border-radius:10px;">
          <tr><td style="padding:14px 16px;font-family:Helvetica,Arial,sans-serif;">
            <div style="font-size:10px;font-weight:bold;letter-spacing:.06em;
                        text-transform:uppercase;color:#6B7280;">{label}</div>
            <div style="font-size:19px;font-weight:bold;color:{color};
                        padding-top:4px;">{value}</div>
          </td></tr>
        </table>"""

    return f"""\
<!DOCTYPE html>
<html lang="es">
<body style="margin:0;padding:0;background:#F4F6F9;">
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" bgcolor="#F4F6F9">
<tr><td align="center" style="padding:32px 16px;">

  <table role="presentation" width="600" cellpadding="0" cellspacing="0" border="0"
         style="max-width:600px;width:100%;background:#FFFFFF;border-radius:12px;
                border:1px solid #E2E6EB;">

    <!-- Header -->
    <tr><td bgcolor="#1B1C20" style="padding:26px 32px;border-radius:12px 12px 0 0;">
      <table role="presentation" cellpadding="0" cellspacing="0" border="0">
        <tr>
          <td bgcolor="#FFFFFF" style="border-radius:8px;padding:7px 12px;">
            <img src="cid:netologo" width="96" alt="neto"
                 style="display:block;width:96px;height:auto;border:0;">
          </td>
          <td style="padding-left:18px;font-family:Helvetica,Arial,sans-serif;color:#FFFFFF;">
            <div style="font-size:16px;font-weight:bold;letter-spacing:-.01em;">
              Diagnóstico Financiero Personal</div>
            <div style="font-size:12px;color:#9CA3AF;padding-top:2px;">
              by Empowered Investor</div>
          </td>
        </tr>
      </table>
    </td></tr>

    <!-- Barra 50-30-20 -->
    <tr><td>{_BAR_5030_20}</td></tr>

    <!-- Cuerpo -->
    <tr><td style="padding:32px;font-family:Helvetica,Arial,sans-serif;color:#1B1C20;">

      <div style="font-size:20px;font-weight:bold;letter-spacing:-.02em;">
        Hola {esc(p["nombre"].split()[0])}, su diagnóstico está listo</div>

      <p style="font-size:14px;line-height:1.6;color:#4B5563;margin:14px 0 24px;">
        Gracias por completar su Diagnóstico Financiero Personal. Adjuntamos el
        reporte completo en Excel con el detalle de sus gastos, ingresos,
        patrimonio y metas. Su asesor recibió una copia y lo contactará para
        coordinar la sesión de asesoría.</p>

      {_html_clasificacion(t)}

      <!-- KPIs -->
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0">
        <tr>
          <td width="49%" valign="top">{kpi("Ingresos / mes", _fmt_crc(t["ingresos"]))}</td>
          <td width="2%">&nbsp;</td>
          <td width="49%" valign="top">{kpi("Gastos / mes", _fmt_crc(t["gastos"]))}</td>
        </tr>
        <tr><td colspan="3" height="10" style="font-size:0;line-height:0;">&nbsp;</td></tr>
        <tr>
          <td width="49%" valign="top">{kpi("Flujo de caja", _fmt_crc(t["flujo"]), flujo_color)}</td>
          <td width="2%">&nbsp;</td>
          <td width="49%" valign="top">{kpi("Patrimonio neto", _fmt_crc(t["patrimonio"]), patri_color)}</td>
        </tr>
      </table>

      {_html_personalidades(p)}

      <!-- Adjunto -->
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0"
             style="margin-top:24px;">
        <tr>
          <td width="4" bgcolor="#3A9C8E" style="border-radius:4px 0 0 4px;font-size:0;">&nbsp;</td>
          <td bgcolor="#F4F6F9" style="padding:14px 16px;border-radius:0 4px 4px 0;
              font-family:Helvetica,Arial,sans-serif;">
            <div style="font-size:13px;font-weight:bold;color:#1B1C20;">
              Reporte adjunto</div>
            <div style="font-size:12.5px;color:#6B7280;padding-top:2px;">
              {esc(archivo)} · generado el {esc(fecha_larga)}</div>
          </td>
        </tr>
      </table>

      <p style="font-size:13px;line-height:1.6;color:#4B5563;margin:24px 0 0;">
        <b>Datos de contacto registrados:</b><br>
        Correo: {esc(p["correo"])}<br>
        Celular: {esc(p["celular"])}</p>

    </td></tr>

    <!-- Footer -->
    <tr><td>{_BAR_5030_20}</td></tr>
    <tr><td bgcolor="#1B1C20" style="padding:20px 32px;border-radius:0 0 12px 12px;">
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0">
        <tr>
          <td style="font-family:Helvetica,Arial,sans-serif;font-size:12px;color:#9CA3AF;">
            <b style="color:#FFFFFF;">neto</b> by Empowered Investor<br>
            <span style="font-size:11px;">50 necesidades · 30 estilo de vida · 20 ahorro</span>
          </td>
          <td align="right" style="font-family:Helvetica,Arial,sans-serif;
              font-size:11px;color:#6B7280;">
            Sus datos no se almacenan<br>en ninguna base de datos.
          </td>
        </tr>
      </table>
    </td></tr>

  </table>

</td></tr>
</table>
</body>
</html>"""


def send_report(p):
    """Envía el reporte al cliente y al asesor. Lanza excepción si falla."""
    host = os.environ.get("SMTP_HOST", "smtp.gmail.com")
    port = int(os.environ.get("SMTP_PORT", "587"))
    user = os.environ.get("SMTP_USER")
    password = os.environ.get("SMTP_PASSWORD")
    if not user or not password:
        raise RuntimeError("SMTP_USER / SMTP_PASSWORD no configurados en .env")

    xlsx = build_excel(p)
    ahora = datetime.now(_CR_TZ)
    fecha = ahora.strftime("%Y-%m-%d")
    archivo = f"diagnostico-financiero-{fecha}.xlsx"

    msg = EmailMessage()
    msg["Subject"] = f"Diagnóstico Financiero — {p['nombre']}"
    msg["From"] = f"neto <{user}>"
    msg["To"] = p["correo"]
    msg["Cc"] = ADVISOR_EMAIL
    msg.set_content(
        f"Hola {p['nombre']},\n\n"
        "Adjuntamos el reporte de su Diagnóstico Financiero Personal. "
        "Su asesor lo revisará y lo contactará para coordinar la sesión de "
        "asesoría.\n"
        + _texto_clasificacion(p) + _texto_personalidades(p) +
        f"\nDatos de contacto registrados:\n"
        f"  • Correo: {p['correo']}\n"
        f"  • Celular: {p['celular']}\n\n"
        "— neto by Empowered Investor\n"
    )

    # Alternativa HTML con la marca neto; el logo va embebido (cid) para que
    # se muestre sin depender de imágenes remotas bloqueadas por el cliente.
    msg.add_alternative(_build_html(p, ahora.strftime("%d/%m/%Y"), archivo),
                        subtype="html")
    with open(_LOGO_PATH, "rb") as f:
        msg.get_payload()[1].add_related(
            f.read(), maintype="image", subtype="png", cid="<netologo>")

    msg.add_attachment(
        xlsx,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=archivo,
    )

    with smtplib.SMTP(host, port, timeout=30) as smtp:
        smtp.starttls()
        smtp.login(user, password)
        smtp.send_message(msg)
