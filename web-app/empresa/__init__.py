import io
import logging
import uuid
from datetime import date, datetime, timedelta
from functools import wraps

import openpyxl
import psycopg2
import psycopg2.extras
from flask import (Blueprint, flash, jsonify, redirect, render_template,
                   request, send_file, session, url_for)
from openpyxl.utils import get_column_letter
from werkzeug.security import check_password_hash, generate_password_hash

import alpaca_client
from crypto import decrypt_secret
from db import (get_connection, get_investment, insert_manual_transaction,
                touch_broker_credentials_used)
from tools import finance
from utils import gen_email_forward, gen_password

empresa_bp = Blueprint('empresa', __name__)


# ── Auth decorators ───────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("empresa.login"))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("empresa.login"))
        if not session.get("business_admin"):
            return redirect(url_for("empresa.transacciones"))
        return f(*args, **kwargs)
    return decorated


# ── Login / Logout ────────────────────────────────────────────────────────────

@empresa_bp.route("/", methods=["GET", "POST"])
def login():
    if "user_id" in session:
        return redirect(url_for("empresa.transacciones"))

    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")

        conn = get_connection()
        try:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """
                    SELECT id, password_hash, business_admin, business_id, client_name,
                           data_privacy_approval, messaging_approval
                    FROM core.clients
                    WHERE username = %s
                    """,
                    (username,),
                )
                user = cur.fetchone()
        finally:
            conn.close()

        if user and user["password_hash"] and check_password_hash(user["password_hash"], password):
            session["user_id"] = str(user["id"])
            session["business_admin"] = bool(user["business_admin"])
            session["business_id"] = str(user["business_id"])
            session["client_name"] = user["client_name"]
            session["consent_ok"] = bool(user["data_privacy_approval"] and user["messaging_approval"])
            return redirect(url_for("empresa.transacciones"))

        error = "Usuario o contraseña incorrectos."

    return render_template("empresa/login.html", error=error)


@empresa_bp.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("empresa.login"))


# ── Consentimiento ────────────────────────────────────────────────────────────

@empresa_bp.route("/consentimiento", methods=["POST"])
@login_required
def consentimiento():
    data_privacy = request.form.get("data_privacy") == "on"
    messaging = request.form.get("messaging") == "on"

    if not (data_privacy and messaging):
        flash("Debe aceptar ambas condiciones para continuar.", "danger")
        return redirect(request.referrer or url_for("empresa.transacciones"))

    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE core.clients
                SET data_privacy_approval = TRUE,
                    messaging_approval    = TRUE,
                    approval_date         = CURRENT_DATE
                WHERE id = %s
                """,
                (session["user_id"],),
            )
        conn.commit()
        session["consent_ok"] = True
    except Exception as e:
        flash(f"Error al guardar consentimiento: {e}", "danger")
    finally:
        conn.close()

    return redirect(request.referrer or url_for("empresa.transacciones"))


# ── Transacciones recientes ───────────────────────────────────────────────────

@empresa_bp.route("/transacciones")
@login_required
def transacciones():
    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT te.merchant_guess,
                       te.amount_guess,
                       te.currency_guess,
                       te.transaction_type_guess,
                       tr.local_date,
                       tn.id          AS notification_id,
                       tn.final_category,
                       tn.final_subcategory
                FROM core.transactions_enriched te
                JOIN core.transactions_raw tr ON te.raw_id = tr.id
                LEFT JOIN core.transactions_classified tc ON tc.raw_id = tr.id
                LEFT JOIN core.transactions_notifications tn ON tn.classified_id = tc.id
                WHERE te.assigned_individual_id = %s
                  AND tr.local_date >= NOW() - INTERVAL '24 hours'
                  AND te.transaction_status NOT IN ('Descartado', 'Duplicado')
                  AND te.transaction_type_guess != 'unknown'
                ORDER BY tr.local_date DESC
                """,
                (session["user_id"],),
            )
            rows = cur.fetchall()

            cur.execute(
                """
                SELECT category, subcategory
                FROM core.categories
                WHERE business_id = %s AND individual_id IS NULL
                ORDER BY category, subcategory NULLS FIRST
                """,
                (session["business_id"],),
            )
            categories = cur.fetchall()
    finally:
        conn.close()

    return render_template("empresa/transacciones.html", rows=rows, categories=categories)


@empresa_bp.route("/transacciones/reclassify", methods=["POST"])
@login_required
def transacciones_reclassify():
    notification_id = request.form.get("notification_id")
    raw_value = request.form.get("category_value", "")
    parts = raw_value.split("|", 1)
    final_category = parts[0].strip() or None
    final_subcategory = parts[1].strip() or None if len(parts) > 1 else None

    if not notification_id or not final_category:
        flash("Datos inválidos.", "danger")
        return redirect(url_for("empresa.transacciones"))

    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE core.transactions_notifications
                SET final_category    = %s,
                    final_subcategory = %s,
                    reclassified_by   = 'user',
                    reclassified_at   = NOW()
                WHERE id = %s AND individual_id = %s
                """,
                (final_category, final_subcategory, notification_id, session["user_id"]),
            )
            updated = cur.rowcount
        conn.commit()
        if updated:
            flash("Reclasificación guardada.", "success")
        else:
            flash("No se encontró la transacción para actualizar.", "warning")
    except Exception as e:
        flash(f"Error al guardar: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for("empresa.transacciones"))


# ── Transacciones pendientes ──────────────────────────────────────────────────

def _count_pending(conn, individual_id):
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT COUNT(*)
            FROM core.transactions_enriched te
            JOIN core.transactions_raw tr ON te.raw_id = tr.id
            WHERE te.assigned_individual_id = %s
              AND te.transaction_type_guess = 'unknown'
              AND te.transaction_approval = 'Aprobada'
              AND te.transaction_status NOT IN ('Descartado', 'Duplicado')
            """,
            (str(individual_id),),
        )
        return cur.fetchone()[0]


@empresa_bp.context_processor
def _inject_pending_count():
    if "user_id" not in session:
        return {}
    try:
        conn = get_connection()
        try:
            return {"pending_count": _count_pending(conn, session["user_id"])}
        finally:
            conn.close()
    except Exception:
        return {"pending_count": 0}


@empresa_bp.context_processor
def _inject_investment_enabled():
    """Controls whether the Inversión tab is shown for this member."""
    if "user_id" not in session:
        return {}
    try:
        conn = get_connection()
        try:
            inv = get_investment(conn, session["user_id"])
            return {"investment_enabled": bool(inv and inv["enabled"])}
        finally:
            conn.close()
    except Exception:
        return {"investment_enabled": False}


@empresa_bp.route("/pendientes")
@login_required
def transacciones_pendientes():
    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT te.id          AS enriched_id,
                       te.merchant_guess,
                       te.amount_guess,
                       te.currency_guess,
                       te.transaction_type_guess,
                       tr.local_date,
                       tn.id          AS notification_id,
                       tn.final_category,
                       tn.final_subcategory
                FROM core.transactions_enriched te
                JOIN core.transactions_raw tr ON te.raw_id = tr.id
                LEFT JOIN core.transactions_classified tc ON tc.raw_id = tr.id
                LEFT JOIN core.transactions_notifications tn ON tn.classified_id = tc.id
                WHERE te.assigned_individual_id = %s
                  AND te.transaction_type_guess = 'unknown'
                  AND te.transaction_approval = 'Aprobada'
                  AND te.transaction_status NOT IN ('Descartado', 'Duplicado')
                ORDER BY tr.local_date DESC
                """,
                (session["user_id"],),
            )
            rows = cur.fetchall()

            cur.execute(
                """
                SELECT category, subcategory
                FROM core.categories
                WHERE business_id = %s AND individual_id IS NULL
                ORDER BY category, subcategory NULLS FIRST
                """,
                (session["business_id"],),
            )
            categories = cur.fetchall()
    finally:
        conn.close()

    return render_template("empresa/pendientes.html", rows=rows, categories=categories)


@empresa_bp.route("/pendientes/save", methods=["POST"])
@login_required
def transacciones_pendientes_save():
    enriched_id = request.form.get("enriched_id")
    notification_id = request.form.get("notification_id") or None
    tipo = request.form.get("tipo", "").strip()
    raw_value = request.form.get("category_value", "")

    if tipo not in ("debito", "credito"):
        flash("Debe seleccionar un tipo (débito o crédito).", "danger")
        return redirect(url_for("empresa.transacciones_pendientes"))
    if not enriched_id:
        flash("Datos inválidos.", "danger")
        return redirect(url_for("empresa.transacciones_pendientes"))

    parts = raw_value.split("|", 1)
    final_category = parts[0].strip() or None
    final_subcategory = parts[1].strip() or None if len(parts) > 1 else None

    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE core.transactions_enriched
                SET transaction_type_guess = %s
                WHERE id = %s AND assigned_individual_id = %s
                """,
                (tipo, enriched_id, session["user_id"]),
            )
            updated = cur.rowcount
            if notification_id and final_category:
                cur.execute(
                    """
                    UPDATE core.transactions_notifications
                    SET final_category    = %s,
                        final_subcategory = %s,
                        reclassified_by   = 'user',
                        reclassified_at   = NOW()
                    WHERE id = %s AND individual_id = %s
                    """,
                    (final_category, final_subcategory, notification_id, session["user_id"]),
                )
        conn.commit()
        if updated:
            flash("Transacción actualizada.", "success")
        else:
            flash("No se encontró la transacción para actualizar.", "warning")
    except Exception as e:
        flash(f"Error al guardar: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for("empresa.transacciones_pendientes"))


# ── Añadir transacciones ──────────────────────────────────────────────────────

_VALID_CURRENCIES = ("CRC", "USD", "EUR")


def _load_categories_empresa(business_id):
    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT category, subcategory
                FROM core.categories
                WHERE business_id = %s AND individual_id IS NULL
                ORDER BY category, subcategory NULLS FIRST
                """,
                (business_id,),
            )
            return cur.fetchall()
    finally:
        conn.close()


@empresa_bp.route("/agregar", methods=["GET", "POST"])
@admin_required
def agregar_transaccion():
    if request.method == "POST":
        merchant = (request.form.get("merchant") or "").strip()
        amount_raw = (request.form.get("amount") or "").strip()
        currency = (request.form.get("currency") or "").strip().upper()
        txn_type = (request.form.get("tipo") or "").strip()
        parts = (request.form.get("category_value") or "").split("|", 1)
        category = parts[0].strip() or None
        subcategory = (parts[1].strip() or None) if len(parts) > 1 else None

        try:
            amount = round(float(amount_raw), 2)
        except (TypeError, ValueError):
            amount = None

        if not merchant:
            flash("Ingrese el comercio.", "danger")
        elif amount is None or amount <= 0:
            flash("Ingrese un monto válido mayor a cero.", "danger")
        elif currency not in _VALID_CURRENCIES:
            flash("Seleccione una moneda válida.", "danger")
        elif txn_type not in ("debito", "credito"):
            flash("Seleccione el tipo (débito o crédito).", "danger")
        elif not category:
            flash("Seleccione una clasificación.", "danger")
        else:
            conn = get_connection()
            try:
                insert_manual_transaction(
                    conn,
                    individual_id=session["user_id"],
                    business_id=session["business_id"],
                    merchant=merchant, amount=amount, currency=currency,
                    txn_type=txn_type, category=category, subcategory=subcategory,
                )
                flash("Transacción agregada correctamente.", "success")
                return redirect(url_for("empresa.transacciones"))
            except Exception as e:
                flash(f"Error al agregar la transacción: {e}", "danger")
            finally:
                conn.close()

    return render_template("empresa/agregar.html",
                           categories=_load_categories_empresa(session["business_id"]))


# ── Dashboard ─────────────────────────────────────────────────────────────────

@empresa_bp.route("/dashboard")
@admin_required
def dashboard():
    return render_template("empresa/dashboard.html")


@empresa_bp.route("/dashboard/general")
@admin_required
def dashboard_general():
    date_from, date_to = finance.resolve_range(request.args.get("range", "ultimo_anio"))
    conn = get_connection()
    try:
        summary = finance.get_income_expense_summary(
            conn, business_id=session["business_id"], date_from=date_from, date_to=date_to)
        top = finance.get_top_spending(
            conn, business_id=session["business_id"], date_from=date_from, date_to=date_to, limit=5)
        categories = finance.list_categories(conn, business_id=session["business_id"])
    finally:
        conn.close()
    return jsonify({
        "summary": summary,
        "top": top,
        "categories": categories,
        "date_from": str(date_from),
        "date_to": str(date_to),
    })


@empresa_bp.route("/dashboard/categoria")
@admin_required
def dashboard_categoria():
    category = request.args.get("cat", "").strip()
    subcategory = request.args.get("sub", "").strip() or None
    if not category:
        return jsonify({"error": "missing category"}), 400
    date_from, date_to = finance.last_12_months_range(date.today())
    conn = get_connection()
    try:
        series = finance.get_monthly_category_spending(
            conn, business_id=session["business_id"], category=category,
            subcategory=subcategory, date_from=date_from, date_to=date_to)
        budget = finance.get_category_budget(
            conn, business_id=session["business_id"], category=category, subcategory=subcategory)
    finally:
        conn.close()
    return jsonify({
        "series": series,
        "budget": budget,
        "category": category,
        "subcategory": subcategory,
    })


# ── Reportes ──────────────────────────────────────────────────────────────────

def _resolve_date_range(filter_type, date_from_str, date_to_str):
    now = datetime.now()
    if filter_type == "mes_anterior":
        last_day = now.replace(day=1) - timedelta(days=1)
        first_day = last_day.replace(day=1)
        return first_day.date(), last_day.date()
    if filter_type == "anio_actual":
        return now.replace(month=1, day=1).date(), now.date()
    if filter_type == "especifica" and date_from_str and date_to_str:
        try:
            return (
                datetime.strptime(date_from_str, "%Y-%m-%d").date(),
                datetime.strptime(date_to_str, "%Y-%m-%d").date(),
            )
        except ValueError:
            pass
    return now.replace(day=1).date(), now.date()


def _fetch_reportes(business_id, date_from, date_to):
    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT c.client_name,
                       tr.message_id,
                       tr.local_date,
                       te.merchant_guess,
                       te.amount_guess,
                       te.currency_guess,
                       te.amount_local,
                       te.currency_local,
                       te.transaction_type_guess,
                       tn.final_category,
                       tn.final_subcategory
                FROM core.transactions_enriched te
                JOIN core.transactions_raw tr ON te.raw_id = tr.id
                JOIN core.clients c ON tr.individual_id = c.id
                LEFT JOIN core.transactions_classified tc ON tc.raw_id = tr.id
                LEFT JOIN core.transactions_notifications tn ON tn.classified_id = tc.id
                WHERE c.business_id = %s
                  AND te.transaction_approval = 'Aprobada'
                  AND te.transaction_status NOT IN ('unknown', 'Descartado', 'Duplicado')
                  AND tr.local_date::date BETWEEN %s AND %s
                ORDER BY tr.local_date DESC
                """,
                (business_id, date_from, date_to),
            )
            return cur.fetchall()
    finally:
        conn.close()


@empresa_bp.route("/reportes")
@admin_required
def reportes():
    filter_type = request.args.get("filter", "mes_actual")
    date_from_str = request.args.get("date_from", "")
    date_to_str = request.args.get("date_to", "")

    date_from, date_to = _resolve_date_range(filter_type, date_from_str, date_to_str)
    rows = _fetch_reportes(session["business_id"], date_from, date_to)

    return render_template(
        "empresa/reportes.html",
        rows=rows,
        filter_type=filter_type,
        date_from=str(date_from),
        date_to=str(date_to),
    )


@empresa_bp.route("/reportes/download")
@admin_required
def reportes_download():
    filter_type = request.args.get("filter", "mes_actual")
    date_from_str = request.args.get("date_from", "")
    date_to_str = request.args.get("date_to", "")

    date_from, date_to = _resolve_date_range(filter_type, date_from_str, date_to_str)
    rows = _fetch_reportes(session["business_id"], date_from, date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transacciones"

    headers = ["A Nombre De", "Message ID", "Fecha", "Comercio", "Tipo", "Moneda", "Monto", "Moneda Local", "Monto Local", "Categoría", "Subcategoría"]
    ws.append(headers)

    for row in rows:
        ws.append([
            row["client_name"],
            row["message_id"],
            row["local_date"].strftime("%Y-%m-%d %H:%M") if row["local_date"] else "",
            row["merchant_guess"] or "",
            row["transaction_type_guess"] or "",
            row["currency_guess"] or "",
            float(row["amount_guess"]) if row["amount_guess"] is not None else "",
            row["currency_local"] or "",
            float(row["amount_local"]) if row["amount_local"] is not None else "",
            row["final_category"] or "",
            row["final_subcategory"] or "",
        ])

    for i, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"transacciones_{date_from}_{date_to}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── Categorias ────────────────────────────────────────────────────────────────

def _parse_budget(value):
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        n = float(s)
        return n if n >= 0 else None
    except (ValueError, TypeError):
        return None


def _upsert_categories(triples, business_id):
    """triples: iterable of (category, subcategory, monthly_budget)."""
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            for category, subcategory, budget in triples:
                budget_value = None if category == "Otros" and subcategory is None else budget
                cur.execute(
                    """
                    INSERT INTO core.categories
                        (id, business_id, individual_id, category, subcategory, monthly_budget)
                    SELECT %s, %s, NULL, %s, %s, %s
                    WHERE NOT EXISTS (
                        SELECT 1 FROM core.categories
                        WHERE business_id = %s
                          AND individual_id IS NULL
                          AND category = %s
                          AND (
                              (subcategory = %s)
                              OR (subcategory IS NULL AND %s IS NULL)
                          )
                    )
                    """,
                    (
                        str(uuid.uuid4()), business_id, category, subcategory, budget_value,
                        business_id, category, subcategory, subcategory,
                    ),
                )
        conn.commit()
    finally:
        conn.close()


@empresa_bp.route("/categorias", methods=["GET", "POST"])
@admin_required
def categorias():
    if request.method == "POST":
        action = request.form.get("action")

        if action == "manual":
            categories = request.form.getlist("category")
            subcategories = request.form.getlist("subcategory")
            budgets = request.form.getlist("monthly_budget")
            triples = [
                (cat.strip(), sub.strip() or None, _parse_budget(bud))
                for cat, sub, bud in zip(
                    categories,
                    subcategories,
                    budgets + [None] * (len(categories) - len(budgets)),
                )
                if cat.strip()
            ]
            if triples:
                _upsert_categories(triples, session["business_id"])
                flash("Categorías guardadas correctamente.", "success")
            else:
                flash("No se ingresaron categorías válidas.", "warning")

        elif action == "upload":
            file = request.files.get("excel_file")
            if not file or not file.filename.lower().endswith((".xlsx", ".xls")):
                flash("Por favor suba un archivo Excel válido (.xlsx).", "danger")
            else:
                try:
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    triples = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        cat = str(row[0]).strip() if row[0] is not None else ""
                        sub = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None
                        bud = _parse_budget(row[2]) if len(row) > 2 else None
                        if cat:
                            triples.append((cat, sub or None, bud))
                    if triples:
                        _upsert_categories(triples, session["business_id"])
                        flash(f"{len(triples)} categoría(s) importada(s) correctamente.", "success")
                    else:
                        flash("El archivo no contiene categorías válidas.", "warning")
                except Exception as e:
                    flash(f"Error al procesar el archivo: {e}", "danger")

        return redirect(url_for("empresa.categorias"))

    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id, category, subcategory, monthly_budget
                FROM core.categories
                WHERE business_id = %s AND individual_id IS NULL
                ORDER BY category, subcategory NULLS FIRST
                """,
                (session["business_id"],),
            )
            existing = cur.fetchall()
    finally:
        conn.close()

    return render_template("empresa/categorias.html", existing=existing)


@empresa_bp.route("/categorias/delete/<categoria_id>", methods=["POST"])
@admin_required
def categorias_delete(categoria_id):
    """Delete a business-level category after reassigning every member's existing
    transactions (and learned rules) to an existing pair or a newly created one.
    All steps run in a single transaction so a failure leaves nothing half-applied."""
    mode = request.form.get("mode")
    business_id = session["business_id"]
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            # 1. Validate the category being deleted
            cur.execute(
                """
                SELECT category, subcategory FROM core.categories
                WHERE id = %s AND business_id = %s AND individual_id IS NULL
                """,
                (categoria_id, business_id),
            )
            row = cur.fetchone()
            if not row:
                flash("Categoría no encontrada.", "danger")
                return redirect(url_for("empresa.categorias"))
            del_cat, del_sub = row[0], row[1]
            if del_cat == "Otros" and del_sub is None:
                flash("La categoría 'Otros' no puede ser eliminada.", "danger")
                return redirect(url_for("empresa.categorias"))

            # 2. Resolve the reassignment target
            if mode == "reassign":
                parts = request.form.get("reassign_value", "").split("|", 1)
                tgt_cat = parts[0].strip() or None
                tgt_sub = (parts[1].strip() or None) if len(parts) > 1 else None
                if not tgt_cat:
                    flash("Debe seleccionar una categoría destino.", "danger")
                    return redirect(url_for("empresa.categorias"))
            elif mode == "new":
                tgt_cat = (request.form.get("new_category") or "").strip() or None
                tgt_sub = (request.form.get("new_subcategory") or "").strip() or None
                if not tgt_cat:
                    flash("Debe ingresar el nombre de la nueva categoría.", "danger")
                    return redirect(url_for("empresa.categorias"))
                cur.execute(
                    """
                    INSERT INTO core.categories
                        (id, business_id, individual_id, category, subcategory, monthly_budget)
                    SELECT %s, %s, NULL, %s, %s, NULL
                    WHERE NOT EXISTS (
                        SELECT 1 FROM core.categories
                        WHERE business_id = %s AND individual_id IS NULL AND category = %s
                          AND ((subcategory = %s) OR (subcategory IS NULL AND %s IS NULL))
                    )
                    """,
                    (str(uuid.uuid4()), business_id, tgt_cat, tgt_sub,
                     business_id, tgt_cat, tgt_sub, tgt_sub),
                )
            else:
                flash("Debe elegir reasignar o crear una nueva categoría.", "danger")
                return redirect(url_for("empresa.categorias"))

            # Target must differ from the pair being deleted
            if tgt_cat == del_cat and (tgt_sub or None) == (del_sub or None):
                flash("La categoría destino debe ser distinta a la que se elimina.", "danger")
                return redirect(url_for("empresa.categorias"))

            # 3. Reassign every member's existing classified transactions
            cur.execute(
                """
                UPDATE core.transactions_notifications
                SET final_category = %s, final_subcategory = %s,
                    reclassified_by = 'user', reclassified_at = NOW()
                WHERE business_id = %s
                  AND final_category = %s
                  AND ((final_subcategory = %s) OR (final_subcategory IS NULL AND %s IS NULL))
                """,
                (tgt_cat, tgt_sub, business_id, del_cat, del_sub, del_sub),
            )
            # 4. Reassign all learned rules in the business (business- and member-level)
            cur.execute(
                """
                UPDATE core.category_rules
                SET category = %s, subcategory = %s, updated_at = NOW()
                WHERE business_id = %s
                  AND category = %s
                  AND ((subcategory = %s) OR (subcategory IS NULL AND %s IS NULL))
                """,
                (tgt_cat, tgt_sub, business_id, del_cat, del_sub, del_sub),
            )
            # 5. Delete the category
            cur.execute(
                """
                DELETE FROM core.categories
                WHERE id = %s AND business_id = %s AND individual_id IS NULL
                """,
                (categoria_id, business_id),
            )
        conn.commit()
        flash("Categoría eliminada y transacciones reasignadas.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Error al eliminar la categoría: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for("empresa.categorias"))


@empresa_bp.route("/categorias/<categoria_id>/budget", methods=["POST"])
@admin_required
def categorias_update_budget(categoria_id):
    budget = _parse_budget(request.form.get("monthly_budget"))
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT category, subcategory FROM core.categories
                WHERE id = %s AND business_id = %s AND individual_id IS NULL
                """,
                (categoria_id, session["business_id"]),
            )
            row = cur.fetchone()
            if not row:
                flash("Categoría no encontrada.", "danger")
                return redirect(url_for("empresa.categorias"))
            if row[0] == "Otros" and row[1] is None:
                flash("La categoría 'Otros' no admite presupuesto.", "danger")
                return redirect(url_for("empresa.categorias"))
            cur.execute(
                """
                UPDATE core.categories SET monthly_budget = %s
                WHERE id = %s AND business_id = %s AND individual_id IS NULL
                """,
                (budget, categoria_id, session["business_id"]),
            )
        conn.commit()
        flash("Presupuesto actualizado.", "success")
    finally:
        conn.close()
    return redirect(url_for("empresa.categorias"))


@empresa_bp.route("/categorias/template")
@admin_required
def categorias_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Categorías"
    ws.append(["categoria", "subcategoria", "presupuesto"])
    ws.append(["Alimentación", "Supermercado", 150000])
    ws.append(["Alimentación", "Restaurantes", 80000])
    ws.append(["Transporte", "Combustible", 50000])
    ws.append(["Transporte", "", ""])
    ws.append(["Entretenimiento", "Cine", ""])

    for i in [1, 2, 3]:
        ws.column_dimensions[get_column_letter(i)].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name="plantilla_categorias.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── Miembros ──────────────────────────────────────────────────────────────────



@empresa_bp.route("/miembros")
@admin_required
def miembros():
    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id, client_name, username, email_address, phone_number,
                       active, email_notification, whatsapp_notification
                FROM core.clients
                WHERE business_id = %s
                ORDER BY client_name
                """,
                (session["business_id"],),
            )
            members = cur.fetchall()
    finally:
        conn.close()

    return render_template("empresa/miembros.html", members=members)


@empresa_bp.route("/miembros/<member_id>/toggle", methods=["POST"])
@admin_required
def miembros_toggle(member_id):
    new_state = request.form.get("active") == "1"
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE core.clients SET active = %s
                WHERE id = %s AND business_id = %s
                """,
                (new_state, member_id, session["business_id"]),
            )
        conn.commit()
    finally:
        conn.close()
    return redirect(url_for("empresa.miembros"))


@empresa_bp.route("/miembros/add", methods=["POST"])
@admin_required
def miembros_add():
    nombre        = request.form.get("nombre", "").strip()
    apellidos     = request.form.get("apellidos", "").strip()
    email_address = request.form.get("email_address", "").strip().lower()
    email_notif   = request.form.get("email_notification") == "1"
    wa_notif      = request.form.get("whatsapp_notification") == "1"
    phone_number  = request.form.get("phone_number", "").strip() or None

    if not nombre or not apellidos or not email_address:
        flash("Nombre, apellidos y correo electrónico son obligatorios.", "warning")
        return redirect(url_for("empresa.miembros"))

    if wa_notif and not phone_number:
        flash("El número de teléfono es obligatorio cuando WhatsApp está activado.", "warning")
        return redirect(url_for("empresa.miembros"))

    client_name   = (nombre + ' ' + apellidos).upper()
    username      = email_address
    password      = gen_password(nombre, apellidos)
    email_forward = gen_email_forward(client_name)

    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO core.clients
                    (id, business_id, business_admin, client_name, email_address,
                     username, password_hash, phone_number, email_forward,
                     active, email_notification, whatsapp_notification)
                VALUES (%s, %s, FALSE, %s, %s, %s, %s, %s, %s, FALSE, %s, %s)
                """,
                (
                    str(uuid.uuid4()),
                    session["business_id"],
                    client_name,
                    email_address,
                    username,
                    generate_password_hash(password),
                    phone_number,
                    email_forward,
                    email_notif,
                    wa_notif,
                ),
            )
        conn.commit()
        flash(f"Miembro '{client_name}' agregado correctamente.", "success")
    except psycopg2.errors.UniqueViolation:
        flash("El correo electrónico ya está registrado. Utilice otro.", "danger")
    except Exception as e:
        flash(f"Error al agregar miembro: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for("empresa.miembros"))


# ── Inversión ─────────────────────────────────────────────────────────────────
#
# Available to every business member (admin and non-admin), gated per member by
# core.client_investment.enabled. The member's Alpaca API credentials are loaded
# by the administrator in the Administración panel and stored encrypted; this
# view only ever decrypts them in memory to READ portfolio data (see the
# security contract in alpaca_client).

@empresa_bp.route("/inversion")
@login_required
def inversion():
    conn = get_connection()
    try:
        inv = get_investment(conn, session["user_id"])

        if not inv or not inv["enabled"]:
            return render_template("empresa/inversion.html", state="disabled")

        # Enabled but credentials not loaded yet by the administrator.
        if not inv["api_key_cipher"] or not inv["api_secret_cipher"]:
            return render_template("empresa/inversion.html", state="pending")

        try:
            aad = str(session["user_id"])
            key_id = decrypt_secret(bytes(inv["api_key_cipher"]), aad=aad)
            api_secret = decrypt_secret(bytes(inv["api_secret_cipher"]), aad=aad)
            portfolio = alpaca_client.get_portfolio_summary_cached(
                key_id, api_secret, str(session["user_id"]))
            touch_broker_credentials_used(conn, session["user_id"])
            return render_template(
                "empresa/inversion.html", state="connected", portfolio=portfolio)
        except Exception:  # noqa: BLE001 - show a recoverable error state
            log.exception(
                "Error consultando el portafolio de Alpaca (client %s)",
                session["user_id"],
            )
            return render_template("empresa/inversion.html", state="error")
    finally:
        conn.close()
