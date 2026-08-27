import io
import logging
import os
import uuid
from datetime import date, datetime, timedelta, timezone
from functools import wraps

import openpyxl
import psycopg2
import psycopg2.extras
from flask import (Blueprint, flash, jsonify, redirect, render_template,
                   request, send_file, session, url_for)
from openpyxl.utils import get_column_letter
from werkzeug.security import check_password_hash, generate_password_hash

import alpaca_client
import paypal_client
from crypto import decrypt_secret
from db import (activate_subscription_from_paypal, apply_discount_code,
                ensure_trial_subscription, get_client_subscription,
                get_connection, get_investment, get_plan_by_paypal_plan_id,
                get_subscription_plan, insert_manual_transaction,
                list_subscription_plans, today_cr,
                touch_broker_credentials_used, update_subscription_from_paypal)
from mailer import enviar_aviso_cambio_contrasena, enviar_link_reset_contrasena
from tools import finance
from utils import (BANK_NOTIFICATION_SENDERS, CLIENT_NOTES_MAX_LEN,
                   clean_client_notes, load_reset_token, make_reset_token,
                   rate_limit_ok, validar_nueva_contrasena)

persona_bp = Blueprint('persona', __name__)

log = logging.getLogger(__name__)

INDIVIDUAL_BIZ_ID = '00000000-0000-0000-0000-000000009999'


# ── Auth decorator ────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session or session.get("app") != "persona":
            return redirect(url_for("persona.login"))
        return f(*args, **kwargs)
    return decorated


# ── Login / Logout ────────────────────────────────────────────────────────────

@persona_bp.route("/", methods=["GET", "POST"])
def login():
    if "user_id" in session and session.get("app") == "persona":
        return redirect(url_for("persona.transacciones"))

    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")

        conn = get_connection()
        try:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """
                    SELECT id, password_hash, business_id, client_name,
                           data_privacy_approval, messaging_approval
                    FROM core.clients
                    WHERE username = %s
                    """,
                    (username,),
                )
                user = cur.fetchone()
        finally:
            conn.close()

        if not user or not user["password_hash"] or not check_password_hash(user["password_hash"], password):
            error = "Usuario o contraseña incorrectos."
        elif str(user["business_id"]) != INDIVIDUAL_BIZ_ID:
            error = "Esta cuenta pertenece a una empresa. Use el portal de Empresa."
        else:
            session.clear()
            session["user_id"] = str(user["id"])
            session["business_id"] = INDIVIDUAL_BIZ_ID
            session["client_name"] = user["client_name"]
            session["app"] = "persona"
            session["consent_ok"] = bool(user["data_privacy_approval"] and user["messaging_approval"])
            return redirect(url_for("persona.transacciones"))

    return render_template("persona/login.html", error=error)


@persona_bp.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("persona.login"))


# ── Contraseña: cambio y restablecimiento ────────────────────────────────────

_RESET_SALT = "pwd-reset-persona"


def _client_ip():
    # Detrás de nginx la IP real viene en X-Forwarded-For
    return (request.headers.get("X-Forwarded-For", request.remote_addr or "?")
            .split(",")[0].strip())


def _actualizar_contrasena(conn, user_id, nueva):
    with conn.cursor() as cur:
        cur.execute(
            """
            UPDATE core.clients
            SET password_hash = %s, password_changed_at = now()
            WHERE id = %s
            """,
            (generate_password_hash(nueva), user_id),
        )
    conn.commit()


@persona_bp.route("/cambiar-contrasena", methods=["GET", "POST"])
@login_required
def cambiar_contrasena():
    error = None
    ok = False
    if request.method == "POST":
        actual = request.form.get("actual", "")
        nueva = request.form.get("nueva", "")
        confirmacion = request.form.get("confirmacion", "")

        conn = get_connection()
        try:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT password_hash, email_address, client_name FROM core.clients WHERE id = %s",
                    (session["user_id"],),
                )
                user = cur.fetchone()

            if not user or not user["password_hash"] or not check_password_hash(user["password_hash"], actual):
                error = "La contraseña actual es incorrecta."
            elif check_password_hash(user["password_hash"], nueva):
                error = "La nueva contraseña no puede ser igual a la actual."
            else:
                error = validar_nueva_contrasena(nueva, confirmacion)

            if error is None:
                _actualizar_contrasena(conn, session["user_id"], nueva)
                ok = True
                try:
                    enviar_aviso_cambio_contrasena(user["email_address"], user["client_name"])
                except Exception as e:
                    log.error(f"Aviso de cambio de contraseña no enviado a {user['email_address']}: {e}")
        finally:
            conn.close()

    return render_template("persona/cambiar_contrasena.html", error=error, ok=ok)


@persona_bp.route("/olvido", methods=["GET", "POST"])
def olvido():
    error = None
    enviado = False
    if request.method == "POST":
        if not rate_limit_ok("olvido-persona", _client_ip()):
            error = "Demasiadas solicitudes desde esta conexión. Intente de nuevo en una hora."
        else:
            username = request.form.get("username", "").strip().lower()
            conn = get_connection()
            try:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    cur.execute(
                        """
                        SELECT id, password_hash, email_address, client_name
                        FROM core.clients
                        WHERE username = %s AND business_id = %s
                        """,
                        (username, INDIVIDUAL_BIZ_ID),
                    )
                    user = cur.fetchone()
            finally:
                conn.close()

            if user and user["email_address"] and user["password_hash"]:
                token = make_reset_token(user["id"], user["password_hash"], _RESET_SALT)
                base = (os.environ.get("WEBAPP_URL") or request.url_root).rstrip("/")
                link = base + url_for("persona.reset_contrasena", token=token)
                try:
                    enviar_link_reset_contrasena(user["email_address"], user["client_name"], link)
                except Exception as e:
                    log.error(f"Correo de restablecimiento no enviado a {user['email_address']}: {e}")
            # Siempre se confirma el envío, exista o no el usuario (evita
            # que el formulario sirva para adivinar usuarios registrados).
            enviado = True

    return render_template("auth/olvido.html", portal="persona", error=error, enviado=enviado)


@persona_bp.route("/reset/<token>", methods=["GET", "POST"])
def reset_contrasena(token):
    data = load_reset_token(token, _RESET_SALT)
    user = None
    if data:
        conn = get_connection()
        try:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    "SELECT id, password_hash, email_address, client_name FROM core.clients WHERE id = %s",
                    (data["uid"],),
                )
                user = cur.fetchone()
        finally:
            conn.close()
        # El token lleva un fragmento del hash vigente al emitirlo: si la
        # contraseña ya cambió (enlace usado u otro reset), el enlace muere.
        if not user or (user["password_hash"] or "")[-16:] != data["ph"]:
            user = None

    if user is None:
        return render_template("auth/reset.html", portal="persona", invalido=True), 400

    error = None
    ok = False
    if request.method == "POST":
        nueva = request.form.get("nueva", "")
        confirmacion = request.form.get("confirmacion", "")
        error = validar_nueva_contrasena(nueva, confirmacion)
        if error is None:
            conn = get_connection()
            try:
                _actualizar_contrasena(conn, user["id"], nueva)
            finally:
                conn.close()
            ok = True
            try:
                enviar_aviso_cambio_contrasena(user["email_address"], user["client_name"])
            except Exception as e:
                log.error(f"Aviso de cambio de contraseña no enviado a {user['email_address']}: {e}")

    return render_template("auth/reset.html", portal="persona", error=error, ok=ok)


# ── Consentimiento ────────────────────────────────────────────────────────────

@persona_bp.route("/consentimiento", methods=["POST"])
@login_required
def consentimiento():
    data_privacy = request.form.get("data_privacy") == "on"
    messaging = request.form.get("messaging") == "on"

    if not (data_privacy and messaging):
        flash("Debe aceptar ambas condiciones para continuar.", "danger")
        return redirect(request.referrer or url_for("persona.transacciones"))

    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE core.clients
                SET data_privacy_approval = TRUE,
                    messaging_approval    = TRUE,
                    approval_date         = CURRENT_DATE
                WHERE id = %s
                """,
                (session["user_id"],),
            )
        conn.commit()
        session["consent_ok"] = True
    except Exception as e:
        flash(f"Error al guardar consentimiento: {e}", "danger")
    finally:
        conn.close()

    return redirect(request.referrer or url_for("persona.transacciones"))


# ── Transacciones recientes ───────────────────────────────────────────────────

@persona_bp.route("/transacciones")
@login_required
def transacciones():
    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT te.id          AS enriched_id,
                       te.merchant_guess,
                       te.amount_guess,
                       te.currency_guess,
                       te.transaction_type_guess,
                       tr.local_date,
                       tn.id          AS notification_id,
                       tn.final_category,
                       tn.final_subcategory,
                       tn.client_notes
                FROM core.transactions_enriched te
                JOIN core.transactions_raw tr ON te.raw_id = tr.id
                LEFT JOIN core.transactions_classified tc ON tc.raw_id = tr.id
                LEFT JOIN core.transactions_notifications tn ON tn.classified_id = tc.id
                WHERE tr.individual_id = %s
                  AND tr.local_date >= NOW() - INTERVAL '24 hours'
                  AND te.transaction_status NOT IN ('Descartado', 'Duplicado')
                  AND te.transaction_type_guess != 'unknown'
                ORDER BY tr.local_date DESC
                """,
                (session["user_id"],),
            )
            rows = cur.fetchall()

            cur.execute(
                """
                SELECT category, subcategory
                FROM core.categories
                WHERE business_id = %s
                  AND (individual_id = %s OR individual_id IS NULL)
                ORDER BY category, subcategory NULLS FIRST
                """,
                (INDIVIDUAL_BIZ_ID, session["user_id"]),
            )
            categories = cur.fetchall()
    finally:
        conn.close()

    return render_template("persona/transacciones.html", rows=rows, categories=categories,
                           notes_max=CLIENT_NOTES_MAX_LEN)


@persona_bp.route("/transacciones/reclassify", methods=["POST"])
@login_required
def transacciones_reclassify():
    notification_id = request.form.get("notification_id")
    raw_value = request.form.get("category_value", "")
    # When invoked from the "Editar transacciones" tab we return there preserving
    # the active date filter. Reclassifying only updates the historical record
    # (final_category); it never touches core.category_rules — the learning signal
    # comes solely from the recent-window flow read by the scheduler.
    origin = request.form.get("origin", "")
    date_from = request.form.get("date_from", "")
    date_to = request.form.get("date_to", "")
    parts = raw_value.split("|", 1)
    final_category = parts[0].strip() or None
    final_subcategory = parts[1].strip() or None if len(parts) > 1 else None
    # La nota se guarda con el mismo botón que la reclasificación. Se toca solo
    # si el campo viajó en el formulario, para no borrarla desde una vista que
    # no la muestre; vacía significa borrarla a propósito.
    notes_submitted = "client_notes" in request.form
    client_notes = clean_client_notes(request.form.get("client_notes"))

    def _back():
        if origin == "editar":
            return redirect(url_for("persona.transacciones_editar",
                                    date_from=date_from, date_to=date_to))
        return redirect(url_for("persona.transacciones"))

    sets, params = [], []
    if final_category:
        sets += ["final_category = %s", "final_subcategory = %s",
                 "reclassified_by = 'user'", "reclassified_at = NOW()"]
        params += [final_category, final_subcategory]
    if notes_submitted:
        sets.append("client_notes = %s")
        params.append(client_notes)

    if not notification_id or not sets:
        flash("Datos inválidos.", "danger")
        return _back()

    conn = get_connection()
    try:
        with conn.cursor() as cur:
            # Los fragmentos de SET los arma este código, nunca el usuario.
            cur.execute(
                "UPDATE core.transactions_notifications SET " + ", ".join(sets) +
                " WHERE id = %s AND individual_id = %s",
                params + [notification_id, session["user_id"]],
            )
            updated = cur.rowcount
        conn.commit()
        if updated:
            flash("Reclasificación guardada." if final_category else "Nota guardada.",
                  "success")
        else:
            flash("No se encontró la transacción para actualizar.", "warning")
    except Exception as e:
        flash(f"Error al guardar: {e}", "danger")
    finally:
        conn.close()

    return _back()


# ── Editar transacciones (descartar) ──────────────────────────────────────────
#
# "Descartar" is a SOFT delete: it sets transactions_enriched.transaction_status
# to 'Descartado'. Every read path (transacciones recientes, pendientes, reportes
# and the finance dashboards/budgets) already excludes that status, so a discarded
# transaction simply disappears everywhere without any physical DELETE across the
# raw → enriched → classified → notifications chain.

def _resolve_editar_range(date_from_str, date_to_str):
    """Defaults to the last week (today − 7 days … today)."""
    today = date.today()
    date_from = today - timedelta(days=7)
    date_to = today
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    return date_from, date_to


@persona_bp.route("/transacciones/editar")
@login_required
def transacciones_editar():
    date_from, date_to = _resolve_editar_range(
        request.args.get("date_from", ""), request.args.get("date_to", ""))
    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT te.id          AS enriched_id,
                       te.merchant_guess,
                       te.amount_guess,
                       te.currency_guess,
                       te.transaction_type_guess,
                       tr.local_date,
                       tn.id          AS notification_id,
                       tn.final_category,
                       tn.final_subcategory,
                       tn.client_notes
                FROM core.transactions_enriched te
                JOIN core.transactions_raw tr ON te.raw_id = tr.id
                LEFT JOIN core.transactions_classified tc ON tc.raw_id = tr.id
                LEFT JOIN core.transactions_notifications tn ON tn.classified_id = tc.id
                WHERE tr.individual_id = %s
                  AND te.transaction_status NOT IN ('Descartado', 'Duplicado')
                  AND te.transaction_type_guess != 'unknown'
                  AND tr.local_date::date BETWEEN %s AND %s
                ORDER BY tr.local_date DESC
                """,
                (session["user_id"], date_from, date_to),
            )
            rows = cur.fetchall()

            cur.execute(
                """
                SELECT category, subcategory
                FROM core.categories
                WHERE business_id = %s
                  AND (individual_id = %s OR individual_id IS NULL)
                ORDER BY category, subcategory NULLS FIRST
                """,
                (INDIVIDUAL_BIZ_ID, session["user_id"]),
            )
            categories = cur.fetchall()
    finally:
        conn.close()

    return render_template(
        "persona/editar.html",
        rows=rows,
        categories=categories,
        date_from=str(date_from),
        date_to=str(date_to),
        notes_max=CLIENT_NOTES_MAX_LEN,
    )


@persona_bp.route("/transacciones/descartar", methods=["POST"])
@login_required
def transacciones_descartar():
    enriched_id = request.form.get("enriched_id")
    # "recientes" returns to Transacciones recientes; otherwise back to Editar
    # preserving the active date filter.
    origin = request.form.get("origin", "")
    date_from = request.form.get("date_from", "")
    date_to = request.form.get("date_to", "")

    def _back():
        if origin == "recientes":
            return redirect(url_for("persona.transacciones"))
        return redirect(url_for("persona.transacciones_editar",
                                date_from=date_from, date_to=date_to))

    if not enriched_id:
        flash("Datos inválidos.", "danger")
        return _back()

    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE core.transactions_enriched
                SET transaction_status = 'Descartado'
                WHERE id = %s AND individual_id = %s
                  AND transaction_status NOT IN ('Descartado', 'Duplicado')
                """,
                (enriched_id, session["user_id"]),
            )
            updated = cur.rowcount
        conn.commit()
        if updated:
            flash("Transacción descartada.", "success")
        else:
            flash("No se encontró la transacción para descartar.", "warning")
    except Exception as e:
        flash(f"Error al descartar: {e}", "danger")
    finally:
        conn.close()

    return _back()


# ── Transacciones pendientes ──────────────────────────────────────────────────

def _count_pending(conn, individual_id):
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT COUNT(*)
            FROM core.transactions_enriched te
            JOIN core.transactions_raw tr ON te.raw_id = tr.id
            WHERE tr.individual_id = %s
              AND te.transaction_type_guess = 'unknown'
              AND te.transaction_approval = 'Aprobada'
              AND te.transaction_status NOT IN ('Descartado', 'Duplicado')
            """,
            (str(individual_id),),
        )
        return cur.fetchone()[0]


@persona_bp.context_processor
def _inject_pending_count():
    if "user_id" not in session or session.get("app") != "persona":
        return {}
    try:
        conn = get_connection()
        try:
            return {"pending_count": _count_pending(conn, session["user_id"])}
        finally:
            conn.close()
    except Exception:
        return {"pending_count": 0}


@persona_bp.route("/pendientes")
@login_required
def transacciones_pendientes():
    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT te.id          AS enriched_id,
                       te.merchant_guess,
                       te.amount_guess,
                       te.currency_guess,
                       te.transaction_type_guess,
                       tr.local_date,
                       tn.id          AS notification_id,
                       tn.final_category,
                       tn.final_subcategory,
                       tn.client_notes
                FROM core.transactions_enriched te
                JOIN core.transactions_raw tr ON te.raw_id = tr.id
                LEFT JOIN core.transactions_classified tc ON tc.raw_id = tr.id
                LEFT JOIN core.transactions_notifications tn ON tn.classified_id = tc.id
                WHERE tr.individual_id = %s
                  AND te.transaction_type_guess = 'unknown'
                  AND te.transaction_approval = 'Aprobada'
                  AND te.transaction_status NOT IN ('Descartado', 'Duplicado')
                ORDER BY tr.local_date DESC
                """,
                (session["user_id"],),
            )
            rows = cur.fetchall()

            cur.execute(
                """
                SELECT category, subcategory
                FROM core.categories
                WHERE business_id = %s
                  AND (individual_id = %s OR individual_id IS NULL)
                ORDER BY category, subcategory NULLS FIRST
                """,
                (INDIVIDUAL_BIZ_ID, session["user_id"]),
            )
            categories = cur.fetchall()
    finally:
        conn.close()

    return render_template("persona/pendientes.html", rows=rows, categories=categories,
                           notes_max=CLIENT_NOTES_MAX_LEN)


@persona_bp.route("/pendientes/save", methods=["POST"])
@login_required
def transacciones_pendientes_save():
    enriched_id = request.form.get("enriched_id")
    notification_id = request.form.get("notification_id") or None
    tipo = request.form.get("tipo", "").strip()
    raw_value = request.form.get("category_value", "")
    notes_submitted = "client_notes" in request.form
    client_notes = clean_client_notes(request.form.get("client_notes"))

    if tipo not in ("debito", "credito"):
        flash("Debe seleccionar un tipo (débito o crédito).", "danger")
        return redirect(url_for("persona.transacciones_pendientes"))
    if not enriched_id:
        flash("Datos inválidos.", "danger")
        return redirect(url_for("persona.transacciones_pendientes"))

    parts = raw_value.split("|", 1)
    final_category = parts[0].strip() or None
    final_subcategory = parts[1].strip() or None if len(parts) > 1 else None

    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE core.transactions_enriched
                SET transaction_type_guess = %s
                WHERE id = %s AND individual_id = %s
                """,
                (tipo, enriched_id, session["user_id"]),
            )
            updated = cur.rowcount
            sets, params = [], []
            if final_category:
                sets += ["final_category = %s", "final_subcategory = %s",
                         "reclassified_by = 'user'", "reclassified_at = NOW()"]
                params += [final_category, final_subcategory]
            if notes_submitted:
                sets.append("client_notes = %s")
                params.append(client_notes)
            if notification_id and sets:
                cur.execute(
                    "UPDATE core.transactions_notifications SET " + ", ".join(sets) +
                    " WHERE id = %s AND individual_id = %s",
                    params + [notification_id, session["user_id"]],
                )
        conn.commit()
        if updated:
            flash("Transacción actualizada.", "success")
        else:
            flash("No se encontró la transacción para actualizar.", "warning")
    except Exception as e:
        flash(f"Error al guardar: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for("persona.transacciones_pendientes"))


# ── Añadir transacciones ──────────────────────────────────────────────────────

_VALID_CURRENCIES = ("CRC", "USD", "EUR")


def _load_categories_persona(user_id):
    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT category, subcategory
                FROM core.categories
                WHERE business_id = %s AND (individual_id = %s OR individual_id IS NULL)
                ORDER BY category, subcategory NULLS FIRST
                """,
                (INDIVIDUAL_BIZ_ID, user_id),
            )
            return cur.fetchall()
    finally:
        conn.close()


@persona_bp.route("/agregar", methods=["GET", "POST"])
@login_required
def agregar_transaccion():
    if request.method == "POST":
        merchant = (request.form.get("merchant") or "").strip()
        amount_raw = (request.form.get("amount") or "").strip()
        currency = (request.form.get("currency") or "").strip().upper()
        txn_type = (request.form.get("tipo") or "").strip()
        txn_date_raw = (request.form.get("txn_date") or "").strip()
        parts = (request.form.get("category_value") or "").split("|", 1)
        category = parts[0].strip() or None
        subcategory = (parts[1].strip() or None) if len(parts) > 1 else None
        client_notes = clean_client_notes(request.form.get("client_notes"))

        try:
            amount = round(float(amount_raw), 2)
        except (TypeError, ValueError):
            amount = None

        # Fecha de la transacción: default hoy (hora CR); nunca futura.
        if txn_date_raw:
            try:
                txn_date = datetime.strptime(txn_date_raw, "%Y-%m-%d").date()
            except ValueError:
                txn_date = None
        else:
            txn_date = today_cr()

        if not merchant:
            flash("Ingrese el comercio.", "danger")
        elif amount is None or amount <= 0:
            flash("Ingrese un monto válido mayor a cero.", "danger")
        elif currency not in _VALID_CURRENCIES:
            flash("Seleccione una moneda válida.", "danger")
        elif txn_type not in ("debito", "credito"):
            flash("Seleccione el tipo (débito o crédito).", "danger")
        elif txn_date is None:
            flash("Ingrese una fecha válida.", "danger")
        elif txn_date > today_cr():
            flash("La fecha de la transacción no puede ser futura.", "danger")
        elif not category:
            flash("Seleccione una clasificación.", "danger")
        else:
            conn = get_connection()
            try:
                insert_manual_transaction(
                    conn,
                    individual_id=session["user_id"],
                    business_id=INDIVIDUAL_BIZ_ID,
                    merchant=merchant, amount=amount, currency=currency,
                    txn_type=txn_type, category=category, subcategory=subcategory,
                    txn_date=txn_date, client_notes=client_notes,
                )
                if txn_date < today_cr():
                    flash(
                        f"Transacción agregada con fecha {txn_date.strftime('%d/%m/%Y')}. "
                        "Por ser de un día anterior no aparece en 'Transacciones recientes' "
                        "(últimas 24 h); la encontrás en Editar transacciones y en Reportes.",
                        "success",
                    )
                else:
                    flash("Transacción agregada correctamente.", "success")
                return redirect(url_for("persona.transacciones"))
            except Exception as e:
                flash(f"Error al agregar la transacción: {e}", "danger")
            finally:
                conn.close()

    return render_template("persona/agregar.html",
                           categories=_load_categories_persona(session["user_id"]),
                           today=str(today_cr()),
                           notes_max=CLIENT_NOTES_MAX_LEN)


# ── Dashboard ─────────────────────────────────────────────────────────────────

@persona_bp.route("/dashboard")
@login_required
def dashboard():
    return render_template("persona/dashboard.html")


@persona_bp.route("/dashboard/general")
@login_required
def dashboard_general():
    date_from, date_to = finance.resolve_range(request.args.get("range", "ultimo_anio"))
    conn = get_connection()
    try:
        summary = finance.get_income_expense_summary(
            conn, individual_id=session["user_id"], date_from=date_from, date_to=date_to)
        top = finance.get_top_spending(
            conn, individual_id=session["user_id"], date_from=date_from, date_to=date_to, limit=5)
        categories = finance.list_categories(conn, individual_id=session["user_id"])
    finally:
        conn.close()
    return jsonify({
        "summary": summary,
        "top": top,
        "categories": categories,
        "date_from": str(date_from),
        "date_to": str(date_to),
    })


@persona_bp.route("/dashboard/categoria")
@login_required
def dashboard_categoria():
    category = request.args.get("cat", "").strip()
    subcategory = request.args.get("sub", "").strip() or None
    if not category:
        return jsonify({"error": "missing category"}), 400
    date_from, date_to = finance.last_12_months_range(date.today())
    conn = get_connection()
    try:
        series = finance.get_monthly_category_spending(
            conn, individual_id=session["user_id"], category=category,
            subcategory=subcategory, date_from=date_from, date_to=date_to)
        budget = finance.get_category_budget(
            conn, individual_id=session["user_id"], category=category, subcategory=subcategory)
    finally:
        conn.close()
    return jsonify({
        "series": series,
        "budget": budget,
        "category": category,
        "subcategory": subcategory,
    })


# ── Inversión ─────────────────────────────────────────────────────────────────
#
# The client's Alpaca API credentials are loaded by the administrator in the
# Administración panel and stored encrypted; this view only ever decrypts them
# in memory to READ portfolio data (see the security contract in alpaca_client).

@persona_bp.route("/inversion")
@login_required
def inversion():
    conn = get_connection()
    try:
        inv = get_investment(conn, session["user_id"])

        # Not an investment client → keep the marketing placeholder.
        if not inv or not inv["enabled"]:
            return render_template("persona/inversion.html", state="disabled")

        # Enabled but credentials not loaded yet by the administrator.
        if not inv["api_key_cipher"] or not inv["api_secret_cipher"]:
            return render_template("persona/inversion.html", state="pending")

        # Configured → read-only portfolio snapshot.
        try:
            aad = str(session["user_id"])
            key_id = decrypt_secret(bytes(inv["api_key_cipher"]), aad=aad)
            api_secret = decrypt_secret(bytes(inv["api_secret_cipher"]), aad=aad)
            portfolio = alpaca_client.get_portfolio_summary_cached(
                key_id, api_secret, str(session["user_id"]))
            touch_broker_credentials_used(conn, session["user_id"])
            return render_template(
                "persona/inversion.html", state="connected", portfolio=portfolio)
        except Exception:  # noqa: BLE001 - show a recoverable error state
            log.exception(
                "Error consultando el portafolio de Alpaca (client %s)",
                session["user_id"],
            )
            return render_template("persona/inversion.html", state="error")
    finally:
        conn.close()


# ── Reportes ──────────────────────────────────────────────────────────────────

def _resolve_date_range(filter_type, date_from_str, date_to_str):
    now = datetime.now()
    if filter_type == "mes_anterior":
        last_day = now.replace(day=1) - timedelta(days=1)
        first_day = last_day.replace(day=1)
        return first_day.date(), last_day.date()
    if filter_type == "anio_actual":
        return now.replace(month=1, day=1).date(), now.date()
    if filter_type == "especifica" and date_from_str and date_to_str:
        try:
            return (
                datetime.strptime(date_from_str, "%Y-%m-%d").date(),
                datetime.strptime(date_to_str, "%Y-%m-%d").date(),
            )
        except ValueError:
            pass
    return now.replace(day=1).date(), now.date()


def _fetch_reportes(individual_id, date_from, date_to):
    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT tr.message_id,
                       tr.local_date,
                       te.merchant_guess,
                       te.amount_guess,
                       te.currency_guess,
                       te.amount_local,
                       te.currency_local,
                       te.transaction_type_guess,
                       tn.final_category,
                       tn.final_subcategory,
                       tn.client_notes
                FROM core.transactions_enriched te
                JOIN core.transactions_raw tr ON te.raw_id = tr.id
                LEFT JOIN core.transactions_classified tc ON tc.raw_id = tr.id
                LEFT JOIN core.transactions_notifications tn ON tn.classified_id = tc.id
                WHERE tr.individual_id = %s
                  AND te.transaction_approval = 'Aprobada'
                  AND te.transaction_status NOT IN ('unknown', 'Descartado', 'Duplicado')
                  AND tr.local_date::date BETWEEN %s AND %s
                ORDER BY tr.local_date DESC
                """,
                (individual_id, date_from, date_to),
            )
            return cur.fetchall()
    finally:
        conn.close()


@persona_bp.route("/reportes")
@login_required
def reportes():
    filter_type = request.args.get("filter", "mes_actual")
    date_from_str = request.args.get("date_from", "")
    date_to_str = request.args.get("date_to", "")

    date_from, date_to = _resolve_date_range(filter_type, date_from_str, date_to_str)
    rows = _fetch_reportes(session["user_id"], date_from, date_to)

    return render_template(
        "persona/reportes.html",
        rows=rows,
        filter_type=filter_type,
        date_from=str(date_from),
        date_to=str(date_to),
    )


@persona_bp.route("/reportes/download")
@login_required
def reportes_download():
    filter_type = request.args.get("filter", "mes_actual")
    date_from_str = request.args.get("date_from", "")
    date_to_str = request.args.get("date_to", "")

    date_from, date_to = _resolve_date_range(filter_type, date_from_str, date_to_str)
    rows = _fetch_reportes(session["user_id"], date_from, date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transacciones"

    headers = ["Message ID", "Fecha", "Comercio", "Tipo", "Moneda", "Monto", "Moneda Local", "Monto Local", "Categoría", "Subcategoría", "Nota"]
    ws.append(headers)

    for row in rows:
        ws.append([
            row["message_id"],
            row["local_date"].strftime("%Y-%m-%d %H:%M") if row["local_date"] else "",
            row["merchant_guess"] or "",
            row["transaction_type_guess"] or "",
            row["currency_guess"] or "",
            float(row["amount_guess"]) if row["amount_guess"] is not None else "",
            row["currency_local"] or "",
            float(row["amount_local"]) if row["amount_local"] is not None else "",
            row["final_category"] or "",
            row["final_subcategory"] or "",
            row["client_notes"] or "",
        ])

    for i, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"transacciones_{date_from}_{date_to}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── Configuración ─────────────────────────────────────────────────────────────
#
# Read-only help page: shows the client's email_forward and provider-specific
# instructions to set up forwarding of bank notifications toward it.

@persona_bp.route("/configuracion")
@login_required
def configuracion():
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT email_forward FROM core.clients WHERE id = %s",
                (session["user_id"],),
            )
            row = cur.fetchone()
    finally:
        conn.close()

    return render_template(
        "persona/configuracion.html",
        email_forward=row[0] if row else None,
        bank_senders=BANK_NOTIFICATION_SENDERS,
    )


# ── Facturación ───────────────────────────────────────────────────────────────
#
# Fase 2: PayPal vivo (API de Suscripciones). El cobro es SIEMPRE en USD
# (PayPal no soporta CRC); la app muestra ₡. El flujo de activación es:
# elegir plan → POST /facturacion/suscribir crea la suscripción en PayPal
# (custom_id = clients.id) → redirect al link de aprobación → PayPal regresa a
# /facturacion/paypal/retorno → paypal_webhook.py confirma/actualiza después.
# La suscripción trial (30 días) se sigue creando perezosa al abrir el tab; si
# el cliente activa un plan durante la prueba, start_time = trial_end difiere
# el primer cobro hasta que la prueba termine.

def _webapp_base_url():
    return os.environ.get("WEBAPP_URL", "https://gastos.empoweredinvestor.trade").rstrip("/")


@persona_bp.route("/facturacion")
@login_required
def facturacion():
    conn = get_connection()
    try:
        ensure_trial_subscription(conn, session["user_id"])
        sub = get_client_subscription(conn, session["user_id"])
        plans = list_subscription_plans(conn, tier="individual")
    finally:
        conn.close()

    paypal_ready = paypal_client.is_configured() and any(p["paypal_plan_id"] for p in plans)
    return render_template("persona/facturacion.html", sub=sub, plans=plans,
                           paypal_ready=paypal_ready)


@persona_bp.route("/facturacion/suscribir", methods=["POST"])
@login_required
def facturacion_suscribir():
    plan_id = (request.form.get("plan_id") or "").strip()
    conn = get_connection()
    try:
        ensure_trial_subscription(conn, session["user_id"])
        sub = get_client_subscription(conn, session["user_id"])
        plan = get_subscription_plan(conn, plan_id) if plan_id else None
    finally:
        conn.close()

    if sub and sub["comp"]:
        flash("Tu cuenta es de cortesía: no necesitás un plan de pago.", "danger")
        return redirect(url_for("persona.facturacion"))
    if sub and sub["status"] == "active" and sub["paypal_subscription_id"]:
        flash("Ya tenés una suscripción activa con PayPal.", "danger")
        return redirect(url_for("persona.facturacion"))
    if (not plan or not plan["active"] or plan["tier"] != "individual"
            or not plan["paypal_plan_id"]):
        flash("Ese plan no está disponible para activar con PayPal.", "danger")
        return redirect(url_for("persona.facturacion"))

    # Prueba gratuita vigente → el primer cobro se difiere al fin de la prueba.
    start_time = None
    now_utc = datetime.now(timezone.utc)
    if (sub and sub["status"] == "trial" and sub["trial_end"]
            and sub["trial_end"] > now_utc):
        start_time = sub["trial_end"].astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    base = _webapp_base_url()
    try:
        _sub_id, approval_url = paypal_client.create_subscription(
            plan["paypal_plan_id"], session["user_id"],
            return_url=f"{base}{url_for('persona.facturacion_paypal_retorno')}",
            cancel_url=f"{base}{url_for('persona.facturacion_paypal_cancelado')}",
            start_time=start_time,
        )
    except Exception as e:
        log.error(f"PayPal: fallo creando suscripción para {session['user_id']}: {e}")
        flash("No se pudo iniciar la suscripción con PayPal. Intentá de nuevo "
              "en unos minutos.", "danger")
        return redirect(url_for("persona.facturacion"))

    return redirect(approval_url)


@persona_bp.route("/facturacion/paypal/retorno")
@login_required
def facturacion_paypal_retorno():
    """Regreso desde la aprobación en PayPal (?subscription_id=I-...).

    Refleja la activación de inmediato para que el cliente no vea el tab
    'en trial' tras aprobar; el webhook ACTIVATED es el que manda a la larga
    (ambos upserts son idempotentes entre sí).
    """
    sub_id = (request.args.get("subscription_id") or "").strip()
    if not sub_id:
        flash("No se recibió la confirmación de PayPal.", "danger")
        return redirect(url_for("persona.facturacion"))

    try:
        psub = paypal_client.get_subscription(sub_id)
    except Exception as e:
        log.error(f"PayPal: fallo leyendo sub {sub_id} en el retorno: {e}")
        flash("Tu suscripción quedó registrada en PayPal; el estado se "
              "reflejará aquí en unos minutos.", "success")
        return redirect(url_for("persona.facturacion"))

    # La suscripción debe ser del cliente logueado (custom_id = clients.id).
    if (psub.get("custom_id") or "") != session["user_id"]:
        log.warning(f"PayPal retorno: sub {sub_id} no pertenece al cliente "
                    f"{session['user_id']} — ignorado")
        flash("No se pudo validar la suscripción de PayPal.", "danger")
        return redirect(url_for("persona.facturacion"))

    if psub.get("status") in ("ACTIVE", "APPROVED"):
        conn = get_connection()
        try:
            plan = get_plan_by_paypal_plan_id(conn, psub.get("plan_id"))
            next_billing = (psub.get("billing_info") or {}).get("next_billing_time")
            activate_subscription_from_paypal(
                conn, session["user_id"], plan["id"] if plan else None,
                sub_id, next_billing)
        finally:
            conn.close()
        flash("¡Listo! Tu plan quedó activo con PayPal.", "success")
    else:
        flash("Tu aprobación quedó registrada; el estado se actualizará en "
              "unos minutos.", "success")
    return redirect(url_for("persona.facturacion"))


@persona_bp.route("/facturacion/paypal/cancelado")
@login_required
def facturacion_paypal_cancelado():
    flash("Activación cancelada. Podés intentarlo cuando querás.", "danger")
    return redirect(url_for("persona.facturacion"))


@persona_bp.route("/facturacion/cancelar", methods=["POST"])
@login_required
def facturacion_cancelar():
    conn = get_connection()
    try:
        sub = get_client_subscription(conn, session["user_id"])
    finally:
        conn.close()

    if not sub or not sub["paypal_subscription_id"] or sub["status"] not in (
            "active", "past_due"):
        flash("No tenés una suscripción de PayPal activa para cancelar.", "danger")
        return redirect(url_for("persona.facturacion"))

    try:
        paypal_client.cancel_subscription(sub["paypal_subscription_id"])
    except Exception as e:
        log.error(f"PayPal: fallo cancelando sub {sub['paypal_subscription_id']}: {e}")
        flash("No se pudo cancelar en PayPal. Intentá de nuevo en unos minutos.",
              "danger")
        return redirect(url_for("persona.facturacion"))

    # El webhook CANCELLED también hará este update (idempotente).
    conn = get_connection()
    try:
        update_subscription_from_paypal(conn, sub["paypal_subscription_id"], "cancelled")
    finally:
        conn.close()
    flash("Suscripción cancelada. No se harán más cobros; el acceso sigue "
          "hasta el fin del período pagado.", "success")
    return redirect(url_for("persona.facturacion"))


@persona_bp.route("/facturacion/codigo", methods=["POST"])
@login_required
def facturacion_codigo():
    code = (request.form.get("code") or "").strip()
    if not code:
        flash("Ingrese un código.", "danger")
        return redirect(url_for("persona.facturacion"))

    conn = get_connection()
    try:
        ok, msg = apply_discount_code(conn, session["user_id"], code)
        flash(msg, "success" if ok else "danger")
    except Exception as e:
        flash(f"Error al aplicar el código: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for("persona.facturacion"))


# ── Categorias ────────────────────────────────────────────────────────────────

def _parse_budget(value):
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        n = float(s)
        return n if n >= 0 else None
    except (ValueError, TypeError):
        return None


def _upsert_categories(triples, individual_id):
    """triples: iterable of (category, subcategory, monthly_budget).
    Inserts are scoped to a specific individual under the sentinel business."""
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            for category, subcategory, budget in triples:
                budget_value = None if category == "Otros" and subcategory is None else budget
                cur.execute(
                    """
                    INSERT INTO core.categories
                        (id, business_id, individual_id, category, subcategory, monthly_budget)
                    SELECT %s, %s, %s, %s, %s, %s
                    WHERE NOT EXISTS (
                        SELECT 1 FROM core.categories
                        WHERE business_id = %s
                          AND individual_id = %s
                          AND category = %s
                          AND (
                              (subcategory = %s)
                              OR (subcategory IS NULL AND %s IS NULL)
                          )
                    )
                    """,
                    (
                        str(uuid.uuid4()), INDIVIDUAL_BIZ_ID, individual_id, category, subcategory, budget_value,
                        INDIVIDUAL_BIZ_ID, individual_id, category, subcategory, subcategory,
                    ),
                )
        conn.commit()
    finally:
        conn.close()


@persona_bp.route("/categorias", methods=["GET", "POST"])
@login_required
def categorias():
    if request.method == "POST":
        action = request.form.get("action")

        if action == "manual":
            categories = request.form.getlist("category")
            subcategories = request.form.getlist("subcategory")
            budgets = request.form.getlist("monthly_budget")
            triples = [
                (cat.strip(), sub.strip() or None, _parse_budget(bud))
                for cat, sub, bud in zip(
                    categories,
                    subcategories,
                    budgets + [None] * (len(categories) - len(budgets)),
                )
                if cat.strip()
            ]
            if triples:
                _upsert_categories(triples, session["user_id"])
                flash("Categorías guardadas correctamente.", "success")
            else:
                flash("No se ingresaron categorías válidas.", "warning")

        elif action == "upload":
            file = request.files.get("excel_file")
            if not file or not file.filename.lower().endswith((".xlsx", ".xls")):
                flash("Por favor suba un archivo Excel válido (.xlsx).", "danger")
            else:
                try:
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    triples = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        cat = str(row[0]).strip() if row[0] is not None else ""
                        sub = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None
                        bud = _parse_budget(row[2]) if len(row) > 2 else None
                        if cat:
                            triples.append((cat, sub or None, bud))
                    if triples:
                        _upsert_categories(triples, session["user_id"])
                        flash(f"{len(triples)} categoría(s) importada(s) correctamente.", "success")
                    else:
                        flash("El archivo no contiene categorías válidas.", "warning")
                except Exception as e:
                    flash(f"Error al procesar el archivo: {e}", "danger")

        return redirect(url_for("persona.categorias"))

    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id, category, subcategory, individual_id, monthly_budget
                FROM core.categories
                WHERE business_id = %s
                  AND (individual_id = %s OR individual_id IS NULL)
                ORDER BY category, subcategory NULLS FIRST
                """,
                (INDIVIDUAL_BIZ_ID, session["user_id"]),
            )
            existing = cur.fetchall()
    finally:
        conn.close()

    return render_template("persona/categorias.html", existing=existing)


@persona_bp.route("/categorias/delete/<categoria_id>", methods=["POST"])
@login_required
def categorias_delete(categoria_id):
    """Delete a category after reassigning its existing transactions (and learned
    rules) to either an existing category/subcategory pair or a newly created one.
    All steps run in a single transaction so a failure leaves nothing half-applied."""
    mode = request.form.get("mode")
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            # 1. Validate the category being deleted
            cur.execute(
                """
                SELECT category, subcategory FROM core.categories
                WHERE id = %s AND business_id = %s AND individual_id = %s
                """,
                (categoria_id, INDIVIDUAL_BIZ_ID, session["user_id"]),
            )
            row = cur.fetchone()
            if not row:
                flash("Categoría no encontrada.", "danger")
                return redirect(url_for("persona.categorias"))
            del_cat, del_sub = row[0], row[1]
            if del_cat == "Otros" and del_sub is None:
                flash("La categoría 'Otros' no puede ser eliminada.", "danger")
                return redirect(url_for("persona.categorias"))

            # 2. Resolve the reassignment target
            if mode == "reassign":
                parts = request.form.get("reassign_value", "").split("|", 1)
                tgt_cat = parts[0].strip() or None
                tgt_sub = (parts[1].strip() or None) if len(parts) > 1 else None
                if not tgt_cat:
                    flash("Debe seleccionar una categoría destino.", "danger")
                    return redirect(url_for("persona.categorias"))
            elif mode == "new":
                tgt_cat = (request.form.get("new_category") or "").strip() or None
                tgt_sub = (request.form.get("new_subcategory") or "").strip() or None
                if not tgt_cat:
                    flash("Debe ingresar el nombre de la nueva categoría.", "danger")
                    return redirect(url_for("persona.categorias"))
                cur.execute(
                    """
                    INSERT INTO core.categories
                        (id, business_id, individual_id, category, subcategory, monthly_budget)
                    SELECT %s, %s, %s, %s, %s, NULL
                    WHERE NOT EXISTS (
                        SELECT 1 FROM core.categories
                        WHERE business_id = %s AND individual_id = %s AND category = %s
                          AND ((subcategory = %s) OR (subcategory IS NULL AND %s IS NULL))
                    )
                    """,
                    (str(uuid.uuid4()), INDIVIDUAL_BIZ_ID, session["user_id"], tgt_cat, tgt_sub,
                     INDIVIDUAL_BIZ_ID, session["user_id"], tgt_cat, tgt_sub, tgt_sub),
                )
            else:
                flash("Debe elegir reasignar o crear una nueva categoría.", "danger")
                return redirect(url_for("persona.categorias"))

            # Target must differ from the pair being deleted
            if tgt_cat == del_cat and (tgt_sub or None) == (del_sub or None):
                flash("La categoría destino debe ser distinta a la que se elimina.", "danger")
                return redirect(url_for("persona.categorias"))

            # 3. Reassign the user's existing classified transactions
            cur.execute(
                """
                UPDATE core.transactions_notifications
                SET final_category = %s, final_subcategory = %s,
                    reclassified_by = 'user', reclassified_at = NOW()
                WHERE individual_id = %s
                  AND final_category = %s
                  AND ((final_subcategory = %s) OR (final_subcategory IS NULL AND %s IS NULL))
                """,
                (tgt_cat, tgt_sub, session["user_id"], del_cat, del_sub, del_sub),
            )
            # 4. Reassign the user's learned rules so the pair isn't resurrected
            cur.execute(
                """
                UPDATE core.category_rules
                SET category = %s, subcategory = %s, updated_at = NOW()
                WHERE business_id = %s AND individual_id = %s
                  AND category = %s
                  AND ((subcategory = %s) OR (subcategory IS NULL AND %s IS NULL))
                """,
                (tgt_cat, tgt_sub, INDIVIDUAL_BIZ_ID, session["user_id"], del_cat, del_sub, del_sub),
            )
            # 5. Delete the category
            cur.execute(
                """
                DELETE FROM core.categories
                WHERE id = %s AND business_id = %s AND individual_id = %s
                """,
                (categoria_id, INDIVIDUAL_BIZ_ID, session["user_id"]),
            )
        conn.commit()
        flash("Categoría eliminada y transacciones reasignadas.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Error al eliminar la categoría: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for("persona.categorias"))


@persona_bp.route("/categorias/<categoria_id>/budget", methods=["POST"])
@login_required
def categorias_update_budget(categoria_id):
    budget = _parse_budget(request.form.get("monthly_budget"))
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT category, subcategory FROM core.categories
                WHERE id = %s AND business_id = %s AND individual_id = %s
                """,
                (categoria_id, INDIVIDUAL_BIZ_ID, session["user_id"]),
            )
            row = cur.fetchone()
            if not row:
                flash("Categoría no encontrada.", "danger")
                return redirect(url_for("persona.categorias"))
            if row[0] == "Otros" and row[1] is None:
                flash("La categoría 'Otros' no admite presupuesto.", "danger")
                return redirect(url_for("persona.categorias"))
            cur.execute(
                """
                UPDATE core.categories SET monthly_budget = %s
                WHERE id = %s AND business_id = %s AND individual_id = %s
                """,
                (budget, categoria_id, INDIVIDUAL_BIZ_ID, session["user_id"]),
            )
        conn.commit()
        flash("Presupuesto actualizado.", "success")
    finally:
        conn.close()
    return redirect(url_for("persona.categorias"))


@persona_bp.route("/categorias/template")
@login_required
def categorias_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Categorías"
    ws.append(["categoria", "subcategoria", "presupuesto"])
    ws.append(["Alimentación", "Supermercado", 150000])
    ws.append(["Alimentación", "Restaurantes", 80000])
    ws.append(["Transporte", "Combustible", 50000])
    ws.append(["Transporte", "", ""])
    ws.append(["Entretenimiento", "Cine", ""])

    for i in [1, 2, 3]:
        ws.column_dimensions[get_column_letter(i)].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name="plantilla_categorias.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
